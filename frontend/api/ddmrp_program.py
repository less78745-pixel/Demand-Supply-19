#!/usr/bin/env python3
"""
DDMRP Program (versi gabungan - 1 file)
=========================================

Program ini menggabungkan semua logika DDMRP ke dalam satu file, dengan dukungan 
eksekusi CLI maupun pemrosesan in-memory untuk API backend (FastAPI):

  1. Baca sheet Raw (data per Cabang/Grup/Category: On Hand, TO, Vessel,
     Forecast, Target per minggu) dan sheet WH (kapasitas gudang per cabang, fix).
  2. Hitung balance mingguan:
       W'      = On Hand + TO(W) + Vessel(W) - Demand(W)
       (W+1)'  = W' + TO(W+1) + Vessel(W+1) - Demand(W+1)   ... dst
     "Demand" = Forecast untuk sheet Hasil Forecast, Target untuk sheet Hasil Target.
  3. Tulis sheet baru: Hasil Forecast, Hasil Target (kolom Perhitungan + Ratio saja,
     TANPA kolom Hasil - Ratio diambil langsung dari kolom Perhitungan).
  4. Tulis sheet Occupancy Forecast & Occupancy Target: total balance per cabang per
     minggu (SUMIF dari kolom Perhitungan) dibagi kapasitas gudang cabang (sheet WH).
  5. Judul kolom minggu dibuat dinamis dari nilai "Week Awal" di sheet WH dan dikonversi 
     ke label periode 4-minggu-per-bulan (1->JAN-1, 2->JAN-2, ..., 8->FEB-4, 9->MAR-1, dst).
  6. Membuat analisa: grafik tren balance & occupancy (matplotlib) + insight otomatis
     (risiko kekurangan stok, risiko over-kapasitas gudang, tren naik/turun signifikan,
     perbandingan skenario Forecast vs Target), dirangkum jadi satu laporan HTML.
"""

import os
import io
import base64
import argparse
from datetime import datetime

from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.worksheet import Worksheet

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker


# ============================================================================
# BAGIAN 1 - CORE: struktur kolom Raw/WH, Week Awal, label periode
# ============================================================================

FIXED_RAW_COLS = 4
ONHAND_COL = 5
FIRST_WEEK_BLOCK_COL = 6
COLS_PER_WEEK = 4

MONTHS = ["JAN", "FEB", "MAR", "APR", "MAY", "JUN",
          "JUL", "AUG", "SEP", "OCT", "NOV", "DEC"]
WEEKS_PER_MONTH = 4


def period_label(week_number: int) -> str:
    """1 -> JAN-1 ; 5 -> FEB-1 ; 8 -> FEB-4 ; 9 -> MAR-1 ; dst. Berputar ke tahun
    berikutnya dengan suffix (Y+1) setelah minggu ke-48."""
    if week_number < 1:
        week_number = 1
    zero_based = week_number - 1
    month_index = zero_based // WEEKS_PER_MONTH
    week_in_month = (zero_based % WEEKS_PER_MONTH) + 1
    year_offset = month_index // 12
    month_name = MONTHS[month_index % 12]
    label = f"{month_name}-{week_in_month}"
    if year_offset > 0:
        label += f" (Y+{year_offset})"
    return label


def read_week_awal(ws_wh: Worksheet, default: int = 1) -> int:
    """Cari sel berisi teks 'Week Awal' (case-insensitive) di sheet WH, ambil nilai
    numerik di sel sebelah kanannya. Contoh: G1='Week Awal', H1=1 -> return 1."""
    for row in ws_wh.iter_rows():
        for cell in row:
            val = cell.value
            if isinstance(val, str) and val.strip().lower() == "week awal":
                next_cell = ws_wh.cell(row=cell.row, column=cell.column + 1)
                if isinstance(next_cell.value, (int, float)):
                    return int(next_cell.value)
    return default


def detect_week_count(ws_raw: Worksheet) -> int:
    max_col = ws_raw.max_column
    remaining = max_col - (FIRST_WEEK_BLOCK_COL - 1)
    n_weeks = remaining // COLS_PER_WEEK
    if n_weeks < 1:
        # If max_column is slightly short or empty, try at least 1 or count non-empty
        return max(1, remaining // COLS_PER_WEEK)
    return n_weeks


def raw_week_cols(w: int):
    """(col_TO, col_Vessel, col_Forecast, col_Target) 1-indexed di sheet Raw, minggu ke-w (0-indexed)."""
    base = FIRST_WEEK_BLOCK_COL + COLS_PER_WEEK * w
    return base, base + 1, base + 2, base + 3


def get_raw_rows(ws_raw: Worksheet):
    rows = []
    for r in range(2, ws_raw.max_row + 1):
        if ws_raw.cell(row=r, column=1).value in (None, "") and ws_raw.cell(row=r, column=2).value in (None, ""):
            continue
        rows.append(r)
    return rows


def get_wh_rows(ws_wh: Worksheet):
    rows = []
    for r in range(2, ws_wh.max_row + 1):
        val = ws_wh.cell(row=r, column=1).value
        cabang = ws_wh.cell(row=r, column=2).value
        if val in (None, "") or cabang in (None, ""):
            continue
        rows.append(r)
    return rows


class RawRecord:
    __slots__ = ("no", "cabang", "grup", "category", "onhand", "to", "vessel", "forecast", "target")

    def __init__(self, no, cabang, grup, category, onhand, to, vessel, forecast, target):
        self.no = no
        self.cabang = str(cabang) if cabang is not None else "Unknown"
        self.grup = str(grup) if grup is not None else "General"
        self.category = str(category) if category is not None else "General"
        self.onhand = float(onhand) if onhand is not None else 0.0
        self.to = [float(x) if x is not None else 0.0 for x in to]
        self.vessel = [float(x) if x is not None else 0.0 for x in vessel]
        self.forecast = [float(x) if x is not None else 0.0 for x in forecast]
        self.target = [float(x) if x is not None else 0.0 for x in target]

    @property
    def label(self):
        return f"{self.cabang}-{self.grup}-{self.category}"


def read_raw_records(ws_raw: Worksheet, n_weeks: int):
    records = []
    for r in get_raw_rows(ws_raw):
        onhand = ws_raw.cell(row=r, column=ONHAND_COL).value or 0
        to, vessel, forecast, target = [], [], [], []
        for w in range(n_weeks):
            c_to, c_vessel, c_forecast, c_target = raw_week_cols(w)
            to.append(ws_raw.cell(row=r, column=c_to).value or 0)
            vessel.append(ws_raw.cell(row=r, column=c_vessel).value or 0)
            forecast.append(ws_raw.cell(row=r, column=c_forecast).value or 0)
            target.append(ws_raw.cell(row=r, column=c_target).value or 0)
        records.append(RawRecord(
            no=ws_raw.cell(row=r, column=1).value,
            cabang=ws_raw.cell(row=r, column=2).value,
            grup=ws_raw.cell(row=r, column=3).value,
            category=ws_raw.cell(row=r, column=4).value,
            onhand=onhand, to=to, vessel=vessel, forecast=forecast, target=target,
        ))
    return records


def read_wh_capacity(ws_wh: Worksheet):
    """Kembalikan list (cabang, kapasitas_total) sesuai urutan baris di sheet WH."""
    result = []
    for r in get_wh_rows(ws_wh):
        cabang = str(ws_wh.cell(row=r, column=2).value)
        existing = ws_wh.cell(row=r, column=3).value or 0
        tambahan = ws_wh.cell(row=r, column=4).value or 0
        try:
            total_cap = float(existing) + float(tambahan)
        except (ValueError, TypeError):
            total_cap = 0.0
        result.append((cabang, total_cap))
    return result


def compute_balance_series(record: RawRecord, n_weeks: int, demand_kind: str):
    """demand_kind: 'forecast' atau 'target'. Kembalikan list balance per minggu."""
    demand = record.forecast if demand_kind == "forecast" else record.target
    balances = []
    prev = 0.0
    for w in range(n_weeks):
        if w == 0:
            bal = record.onhand + record.to[0] + record.vessel[0] - demand[0]
        else:
            bal = prev + record.to[w] + record.vessel[w] - demand[w]
        balances.append(bal)
        prev = bal
    return balances


def compute_ratio_series(balances, record: RawRecord, n_weeks: int, demand_kind: str):
    """Ratio minggu w = balance(w) / demand(w+1). Minggu terakhir -> None."""
    demand = record.forecast if demand_kind == "forecast" else record.target
    ratios = []
    for w in range(n_weeks):
        if w + 1 < n_weeks and demand[w + 1]:
            ratios.append(balances[w] / demand[w + 1])
        else:
            ratios.append(None)
    return ratios


def compute_occupancy(records, balances_by_record, wh_capacity):
    if not balances_by_record:
        return {}
    n_weeks = len(next(iter(balances_by_record.values())))
    occupancy = {}
    for cabang, capacity in wh_capacity:
        totals = [0.0] * n_weeks
        for rec in records:
            if rec.cabang == cabang:
                bals = balances_by_record[id(rec)]
                for w in range(n_weeks):
                    totals[w] += bals[w]
        occupancy[cabang] = [t / capacity if capacity else None for t in totals]
    return occupancy


def build_period_labels(week_awal: int, n_weeks: int):
    return [period_label(week_awal + w) for w in range(n_weeks)]


# ============================================================================
# BAGIAN 2 - GENERATOR EXCEL (Hasil Forecast / Hasil Target / Occupancy)
# ============================================================================

FILL_JUDUL = PatternFill("solid", fgColor="D9D9D9")
FILL_PERHITUNGAN = PatternFill("solid", fgColor="FCE4D6")
FILL_RATIO = PatternFill("solid", fgColor="CFE2F3")
FILL_HEADER2 = PatternFill("solid", fgColor="F2F2F2")
BOLD = Font(bold=True)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_header_cell(ws, row, col, text, fill=None):
    c = ws.cell(row=row, column=col, value=text)
    c.font = BOLD
    c.alignment = CENTER
    c.border = BORDER
    if fill:
        c.fill = fill
    return c


def build_hasil_sheet(wb, sheet_name, ws_raw, raw_rows, n_weeks, demand_kind, period_labels):
    assert demand_kind in ("forecast", "target")

    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)

    col_perhitungan_start = FIXED_RAW_COLS + 1          # 5
    col_ratio_start = col_perhitungan_start + n_weeks     # 5+n
    total_cols = col_ratio_start + n_weeks - 1              # 4+2n

    for c in range(1, FIXED_RAW_COLS + 1):
        style_header_cell(ws, 1, c, "Judul", FILL_JUDUL)
    for w in range(n_weeks):
        style_header_cell(ws, 1, col_perhitungan_start + w, "Perhitungan", FILL_PERHITUNGAN)
    for w in range(n_weeks):
        style_header_cell(ws, 1, col_ratio_start + w, "Ratio", FILL_RATIO)

    fixed_titles = ["No", "Cabang", "Grup", "Category"]
    for i, t in enumerate(fixed_titles, start=1):
        style_header_cell(ws, 2, i, t, FILL_HEADER2)
    for w in range(n_weeks):
        style_header_cell(ws, 2, col_perhitungan_start + w, period_labels[w], FILL_HEADER2)
    for w in range(n_weeks):
        style_header_cell(ws, 2, col_ratio_start + w, period_labels[w], FILL_HEADER2)

    out_row = 3
    for raw_row in raw_rows:
        ws.cell(row=out_row, column=1, value=f"=Raw!A{raw_row}")
        ws.cell(row=out_row, column=2, value=f"=Raw!B{raw_row}")
        ws.cell(row=out_row, column=3, value=f"=Raw!C{raw_row}")
        ws.cell(row=out_row, column=4, value=f"=Raw!D{raw_row}")

        perhitungan_col_letters = []
        for w in range(n_weeks):
            col_to, col_vessel, col_forecast, col_target = raw_week_cols(w)
            col_demand = col_forecast if demand_kind == "forecast" else col_target
            L_to = get_column_letter(col_to)
            L_vessel = get_column_letter(col_vessel)
            L_demand = get_column_letter(col_demand)

            out_col = col_perhitungan_start + w
            if w == 0:
                formula = f"=SUM(Raw!E{raw_row},Raw!{L_to}{raw_row},Raw!{L_vessel}{raw_row})-Raw!{L_demand}{raw_row}"
            else:
                prev_letter = perhitungan_col_letters[w - 1]
                formula = f"=SUM({prev_letter}{out_row},Raw!{L_to}{raw_row},Raw!{L_vessel}{raw_row})-Raw!{L_demand}{raw_row}"
            ws.cell(row=out_row, column=out_col, value=formula)
            perhitungan_col_letters.append(get_column_letter(out_col))

        for w in range(n_weeks):
            out_col = col_ratio_start + w
            if w + 1 < n_weeks:
                perhitungan_letter = perhitungan_col_letters[w]
                _, _, col_forecast_next, col_target_next = raw_week_cols(w + 1)
                col_demand_next = col_forecast_next if demand_kind == "forecast" else col_target_next
                L_demand_next = get_column_letter(col_demand_next)
                formula = (
                    f'=IFERROR(IF(OR(Raw!{L_demand_next}{raw_row}="",Raw!{L_demand_next}{raw_row}=0),"",'
                    f'{perhitungan_letter}{out_row}/Raw!{L_demand_next}{raw_row}),"")'
                )
                ws.cell(row=out_row, column=out_col, value=formula)
                ws.cell(row=out_row, column=out_col).number_format = "0.0%"
            else:
                ws.cell(row=out_row, column=out_col, value=None)

        for c in range(1, total_cols + 1):
            ws.cell(row=out_row, column=c).border = BORDER

        out_row += 1

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 10
    for c in range(col_perhitungan_start, total_cols + 1):
        ws.column_dimensions[get_column_letter(c)].width = 12

    ws.freeze_panes = "E3"
    return ws, col_perhitungan_start, out_row - 1


def build_occupancy_sheet(wb, sheet_name, ws_wh, wh_rows, hasil_sheet_name,
                            perhitungan_col_start, n_weeks, hasil_last_row, period_labels):
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)

    titles = ["No", "Cabang"] + period_labels
    for i, t in enumerate(titles, start=1):
        style_header_cell(ws, 1, i, t, FILL_HEADER2)

    out_row = 2
    for wh_row in wh_rows:
        ws.cell(row=out_row, column=1, value=f"=WH!A{wh_row}")
        ws.cell(row=out_row, column=2, value=f"=WH!B{wh_row}")

        for w in range(n_weeks):
            col_letter = get_column_letter(perhitungan_col_start + w)
            out_col = 3 + w
            formula = (
                f"=IFERROR(IF(WH!$E{wh_row}=0,0,SUMIF('{hasil_sheet_name}'!$B$3:$B${hasil_last_row},"
                f"$B{out_row},"
                f"'{hasil_sheet_name}'!${col_letter}$3:${col_letter}${hasil_last_row})"
                f"/WH!$E{wh_row}),0)"
            )
            ws.cell(row=out_row, column=out_col, value=formula)
            ws.cell(row=out_row, column=out_col).number_format = "0.0%"

        for c in range(1, 3 + n_weeks):
            ws.cell(row=out_row, column=c).border = BORDER

        out_row += 1

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 14
    for w in range(n_weeks):
        ws.column_dimensions[get_column_letter(3 + w)].width = 12
    ws.freeze_panes = "C2"
    return ws


def generate_excel_workbook(wb: Workbook):
    """Memodifikasi Workbook in-place dengan menambahkan sheet Hasil & Occupancy."""
    if "Raw" not in wb.sheetnames:
        raise ValueError("Sheet 'Raw' tidak ditemukan di file input.")
    if "WH" not in wb.sheetnames:
        raise ValueError("Sheet 'WH' tidak ditemukan di file input.")

    ws_raw = wb["Raw"]
    ws_wh = wb["WH"]

    n_weeks = detect_week_count(ws_raw)
    raw_rows = get_raw_rows(ws_raw)
    wh_rows = get_wh_rows(ws_wh)
    week_awal = read_week_awal(ws_wh, default=1)
    period_labels = build_period_labels(week_awal, n_weeks)

    for r in wh_rows:
        ws_wh.cell(row=r, column=5, value=f"=C{r}+D{r}")

    ws_hf, hf_perhitungan_start, hf_last_row = build_hasil_sheet(
        wb, "Hasil Forecast", ws_raw, raw_rows, n_weeks, "forecast", period_labels
    )
    ws_ht, ht_perhitungan_start, ht_last_row = build_hasil_sheet(
        wb, "Hasil Target", ws_raw, raw_rows, n_weeks, "target", period_labels
    )

    build_occupancy_sheet(
        wb, "Occupancy Forecast", ws_wh, wh_rows, "Hasil Forecast",
        hf_perhitungan_start, n_weeks, hf_last_row, period_labels
    )
    build_occupancy_sheet(
        wb, "Occupancy Target", ws_wh, wh_rows, "Hasil Target",
        ht_perhitungan_start, n_weeks, ht_last_row, period_labels
    )

    order = ["Raw", "WH", "Hasil Forecast", "Hasil Target", "Occupancy Forecast", "Occupancy Target"]
    wb._sheets.sort(key=lambda s: order.index(s.title) if s.title in order else 999)
    return wb, period_labels, week_awal, n_weeks


def generate_excel(input_path, output_path):
    wb = load_workbook(input_path)
    wb, period_labels, week_awal, n_weeks = generate_excel_workbook(wb)
    wb.save(output_path)
    print(f"[Excel] Selesai. File hasil disimpan di: {output_path}")


# ============================================================================
# BAGIAN 3 - ANALISA & GRAFIK (matplotlib) + INSIGHT + LAPORAN HTML
# ============================================================================

CATEGORICAL = ["#2a78d6", "#eb6834", "#1baf7a", "#eda100",
               "#e87ba4", "#008300", "#4a3aa7", "#e34948"]
SURFACE = "#fcfcfb"
INK_PRIMARY = "#0b0b0b"
INK_SECONDARY = "#52514e"
INK_MUTED = "#898781"
GRIDLINE = "#e1e0d9"
BASELINE = "#c3c2b7"
STATUS_CRITICAL = "#d03b3b"

plt.rcParams.update({
    "figure.facecolor": SURFACE,
    "axes.facecolor": SURFACE,
    "axes.edgecolor": BASELINE,
    "axes.labelcolor": INK_SECONDARY,
    "text.color": INK_PRIMARY,
    "xtick.color": INK_MUTED,
    "ytick.color": INK_MUTED,
    "font.family": "sans-serif",
    "font.size": 11,
    "axes.grid": True,
    "grid.color": GRIDLINE,
    "grid.linewidth": 0.8,
    "axes.spines.top": False,
    "axes.spines.right": False,
})


def _style_axes(ax):
    ax.grid(axis="y", which="major")
    ax.grid(axis="x", visible=False)
    ax.spines["left"].set_color(BASELINE)
    ax.spines["bottom"].set_color(BASELINE)


def plot_balance_chart(records, balances_by_record, period_labels, title, filename=None):
    fig, ax = plt.subplots(figsize=(9.5, 5))
    _style_axes(ax)
    for i, rec in enumerate(records):
        color = CATEGORICAL[i % len(CATEGORICAL)]
        y = balances_by_record[id(rec)]
        ax.plot(period_labels, y, marker="o", markersize=5, linewidth=2,
                color=color, label=rec.label)
    ax.axhline(0, color=STATUS_CRITICAL, linewidth=1, linestyle="--", alpha=0.6)
    ax.set_ylabel("Balance (unit)")
    ax.set_title(title, fontsize=13, fontweight="bold", color=INK_PRIMARY, loc="left")
    ax.legend(loc="upper left", bbox_to_anchor=(1.0, 1.0), frameon=False, fontsize=9)
    fig.subplots_adjust(right=0.78)
    
    if filename:
        fig.savefig(filename, dpi=150, facecolor=SURFACE)
        plt.close(fig)
        return filename
    else:
        buf = io.BytesIO()
        fig.savefig(buf, format="png", dpi=150, facecolor=SURFACE)
        plt.close(fig)
        buf.seek(0)
        return base64.b64encode(buf.read()).decode("ascii")


def plot_occupancy_chart(occupancy, period_labels, title, filename=None):
    fig, ax = plt.subplots(figsize=(8.5, 5))
    _style_axes(ax)
    cabang_list = list(occupancy.keys())
    for i, cabang in enumerate(cabang_list):
        color = CATEGORICAL[i % len(CATEGORICAL)]
        y = [v * 100 if v is not None else None for v in occupancy[cabang]]
        ax.plot(period_labels, y, marker="o", markersize=6, linewidth=2.2,
                color=color, label=cabang)
    ax.axhline(100, color=STATUS_CRITICAL, linewidth=1.4, linestyle="--")
    ax.text(0.01, 100, " Kapasitas 100% ", color=STATUS_CRITICAL, fontsize=9,
             va="bottom", ha="left", transform=ax.get_yaxis_transform())
    ax.yaxis.set_major_formatter(mticker.PercentFormatter())
    ax.set_ylabel("Occupancy (%)")
    ax.set_title(title, fontsize=13, fontweight="bold", color=INK_PRIMARY, loc="left")
    ax.legend(loc="upper left", bbox_to_anchor=(1.0, 1.0), frameon=False, fontsize=9)
    fig.subplots_adjust(right=0.8)
    
    if filename:
        fig.savefig(filename, dpi=150, facecolor=SURFACE)
        plt.close(fig)
        return filename
    else:
        buf = io.BytesIO()
        fig.savefig(buf, format="png", dpi=150, facecolor=SURFACE)
        plt.close(fig)
        buf.seek(0)
        return base64.b64encode(buf.read()).decode("ascii")


def generate_insights(records, bal_f, bal_t, ratio_f, ratio_t,
                       occ_f, occ_t, period_labels, n_weeks):
    insights = []

    for rec in records:
        rf = ratio_f[id(rec)]
        for w, r in enumerate(rf):
            if r is not None and r < 1:
                insights.append(
                    f"RISIKO KEKURANGAN (Forecast) - {rec.label}: pada periode {period_labels[w]}, "
                    f"balance hanya {r*100:.0f}% dari kebutuhan forecast periode {period_labels[w+1]}. "
                    f"Perlu tambahan TO/Vessel masuk sebelum periode {period_labels[w+1]}."
                )
    for rec in records:
        rt = ratio_t[id(rec)]
        for w, r in enumerate(rt):
            if r is not None and r < 1:
                insights.append(
                    f"RISIKO KEKURANGAN (Target) - {rec.label}: pada periode {period_labels[w]}, "
                    f"balance hanya {r*100:.0f}% dari target periode {period_labels[w+1]}."
                )

    for cabang, series in occ_f.items():
        for w, v in enumerate(series):
            if v is not None and v > 1:
                insights.append(
                    f"RISIKO OVER KAPASITAS GUDANG (Forecast) - Cabang {cabang} pada periode "
                    f"{period_labels[w]}: occupancy {v*100:.0f}% dari kapasitas gudang "
                    f"(kelebihan {(v-1)*100:.0f} poin persen)."
                )
    for cabang, series in occ_t.items():
        for w, v in enumerate(series):
            if v is not None and v > 1:
                insights.append(
                    f"RISIKO OVER KAPASITAS GUDANG (Target) - Cabang {cabang} pada periode "
                    f"{period_labels[w]}: occupancy {v*100:.0f}% dari kapasitas gudang."
                )

    for rec in records:
        yf = bal_f[id(rec)]
        delta = yf[-1] - yf[0]
        pct = (delta / abs(yf[0]) * 100) if yf[0] else None
        arah = "naik" if delta > 0 else ("turun" if delta < 0 else "stabil")
        if pct is not None and abs(pct) >= 30:
            insights.append(
                f"TREN - {rec.label} (Forecast): balance {arah} signifikan dari "
                f"{yf[0]:.1f} ({period_labels[0]}) menjadi {yf[-1]:.1f} ({period_labels[-1]}), "
                f"perubahan {pct:+.0f}%. "
                + ("Cek risiko overstock." if delta > 0 else "Cek risiko kekurangan stok.")
            )

    for rec in records:
        yf = bal_f[id(rec)][-1]
        yt = bal_t[id(rec)][-1]
        gap = yf - yt
        if abs(gap) > 0.5:
            lebih_besar = "Forecast" if gap > 0 else "Target"
            insights.append(
                f"PERBANDINGAN SKENARIO - {rec.label}: pada periode {period_labels[-1]}, balance "
                f"skenario {lebih_besar} lebih tinggi (selisih {abs(gap):.1f} unit) dibanding skenario lainnya. "
                f"Artinya asumsi demand {lebih_besar.lower()} menghasilkan buffer yang lebih besar."
            )

    if not insights:
        insights.append(
            "Tidak ditemukan indikasi risiko kekurangan stok atau over-kapasitas gudang "
            "pada rentang periode yang dianalisis."
        )
    return insights


def _img_b64(path):
    if os.path.exists(path):
        with open(path, "rb") as f:
            return base64.b64encode(f.read()).decode("ascii")
    return path  # assume already base64 string


def build_html_report_string(charts_data, insights, period_labels, week_awal):
    """charts_data list tuple (title, path_or_base64)."""
    imgs_html = "".join(
        f'<div class="chart"><h3>{title}</h3>'
        f'<img src="data:image/png;base64,{_img_b64(img_src)}" alt="{title}"/></div>'
        for title, img_src in charts_data
    )
    insight_items = "".join(f"<li>{i}</li>" for i in insights)

    html = f"""<!DOCTYPE html>
<html lang="id">
<head>
<meta charset="utf-8"/>
<title>DDMRP Analysis Report</title>
<style>
  body {{ font-family: system-ui, -apple-system, "Segoe UI", sans-serif;
          background:#f9f9f7; color:#0b0b0b; margin:0; padding:32px; }}
  h1 {{ font-size: 22px; margin-bottom:4px; }}
  .sub {{ color:#52514e; margin-bottom:28px; }}
  .chart {{ background:#fcfcfb; border:1px solid #e1e0d9; border-radius:10px;
            padding:16px; margin-bottom:24px; }}
  .chart h3 {{ margin-top:0; font-size:15px; }}
  .chart img {{ max-width:100%; display:block; }}
  .insights {{ background:#fcfcfb; border:1px solid #e1e0d9; border-radius:10px; padding:20px 28px; }}
  .insights li {{ margin-bottom:10px; line-height:1.5; }}
  .grid {{ display:grid; grid-template-columns: 1fr 1fr; gap:20px; }}
  @media (max-width: 900px) {{ .grid {{ grid-template-columns: 1fr; }} }}
</style>
</head>
<body>
  <h1>DDMRP Balance &amp; Occupancy - Analysis Report</h1>
  <div class="sub">Periode: {period_labels[0]} - {period_labels[-1]} (Week Awal = {week_awal})</div>
  <div class="grid">
    {imgs_html}
  </div>
  <h2>Insight &amp; Rekomendasi</h2>
  <div class="insights">
    <ul>
      {insight_items}
    </ul>
  </div>
</body>
</html>
"""
    return html


def build_html_report(chart_paths, insights, period_labels, week_awal, output_path):
    html = build_html_report_string(chart_paths, insights, period_labels, week_awal)
    with open(output_path, "w", encoding="utf-8") as f:
        f.write(html)


def generate_analysis(input_path, output_html, charts_dir):
    os.makedirs(charts_dir, exist_ok=True)
    wb = load_workbook(input_path, data_only=True)
    ws_raw = wb["Raw"]
    ws_wh = wb["WH"]

    n_weeks = detect_week_count(ws_raw)
    week_awal = read_week_awal(ws_wh, default=1)
    period_labels = [period_label(week_awal + w) for w in range(n_weeks)]

    records = read_raw_records(ws_raw, n_weeks)
    wh_capacity = read_wh_capacity(ws_wh)

    bal_f, bal_t, ratio_f, ratio_t = {}, {}, {}, {}
    for rec in records:
        bf = compute_balance_series(rec, n_weeks, "forecast")
        bt = compute_balance_series(rec, n_weeks, "target")
        bal_f[id(rec)] = bf
        bal_t[id(rec)] = bt
        ratio_f[id(rec)] = compute_ratio_series(bf, rec, n_weeks, "forecast")
        ratio_t[id(rec)] = compute_ratio_series(bt, rec, n_weeks, "target")

    occ_f = compute_occupancy(records, bal_f, wh_capacity)
    occ_t = compute_occupancy(records, bal_t, wh_capacity)

    p_bal_f = os.path.join(charts_dir, "balance_forecast.png")
    p_bal_t = os.path.join(charts_dir, "balance_target.png")
    p_occ_f = os.path.join(charts_dir, "occupancy_forecast.png")
    p_occ_t = os.path.join(charts_dir, "occupancy_target.png")

    plot_balance_chart(records, bal_f, period_labels,
                        "Tren Balance per Cabang/Grup/Category - Skenario Forecast", p_bal_f)
    plot_balance_chart(records, bal_t, period_labels,
                        "Tren Balance per Cabang/Grup/Category - Skenario Target", p_bal_t)
    plot_occupancy_chart(occ_f, period_labels,
                          "Occupancy Gudang per Cabang - Skenario Forecast", p_occ_f)
    plot_occupancy_chart(occ_t, period_labels,
                          "Occupancy Gudang per Cabang - Skenario Target", p_occ_t)

    insights = generate_insights(records, bal_f, bal_t, ratio_f, ratio_t,
                                  occ_f, occ_t, period_labels, n_weeks)

    build_html_report(
        [
            ("Balance - Skenario Forecast", p_bal_f),
            ("Balance - Skenario Target", p_bal_t),
            ("Occupancy - Skenario Forecast", p_occ_f),
            ("Occupancy - Skenario Target", p_occ_t),
        ],
        insights, period_labels, week_awal, output_html,
    )

    print(f"[Analisa] Periode dianalisis: {', '.join(period_labels)}")
    print(f"[Analisa] Laporan HTML: {output_html}")
    print(f"[Analisa] Grafik PNG  : {charts_dir}/")
    print("\n=== INSIGHT ===")
    for i in insights:
        print(f"- {i}")


def process_ddmrp_in_memory(file_bytes: bytes) -> dict:
    """
    Eksekusi penuh logika DDMRP secara in-memory untuk API backend.
    Mengeluarkan struktur hasil yang 100% kompatibel dengan frontend Occupancy & Inventory.
    """
    wb_data = load_workbook(io.BytesIO(file_bytes), data_only=True)
    if "Raw" not in wb_data.sheetnames or "WH" not in wb_data.sheetnames:
        raise ValueError("File Excel harus memiliki sheet 'Raw' dan 'WH'")

    ws_raw = wb_data["Raw"]
    ws_wh = wb_data["WH"]

    n_weeks = detect_week_count(ws_raw)
    week_awal = read_week_awal(ws_wh, default=1)
    period_labels = build_period_labels(week_awal, n_weeks)

    records = read_raw_records(ws_raw, n_weeks)
    wh_capacity = read_wh_capacity(ws_wh)

    bal_f, bal_t, ratio_f, ratio_t = {}, {}, {}, {}
    for rec in records:
        bf = compute_balance_series(rec, n_weeks, "forecast")
        bt = compute_balance_series(rec, n_weeks, "target")
        bal_f[id(rec)] = bf
        bal_t[id(rec)] = bt
        ratio_f[id(rec)] = compute_ratio_series(bf, rec, n_weeks, "forecast")
        ratio_t[id(rec)] = compute_ratio_series(bt, rec, n_weeks, "target")

    occ_f = compute_occupancy(records, bal_f, wh_capacity)
    occ_t = compute_occupancy(records, bal_t, wh_capacity)

    # Generate charts as base64
    b64_bal_f = plot_balance_chart(records, bal_f, period_labels, "Tren Balance - Skenario Forecast")
    b64_bal_t = plot_balance_chart(records, bal_t, period_labels, "Tren Balance - Skenario Target")
    b64_occ_f = plot_occupancy_chart(occ_f, period_labels, "Occupancy Gudang - Skenario Forecast")
    b64_occ_t = plot_occupancy_chart(occ_t, period_labels, "Occupancy Gudang - Skenario Target")

    insights = generate_insights(records, bal_f, bal_t, ratio_f, ratio_t, occ_f, occ_t, period_labels, n_weeks)

    html_report = build_html_report_string([
        ("Balance - Skenario Forecast", b64_bal_f),
        ("Balance - Skenario Target", b64_bal_t),
        ("Occupancy - Skenario Forecast", b64_occ_f),
        ("Occupancy - Skenario Target", b64_occ_t),
    ], insights, period_labels, week_awal)

    # Generate formulas Excel Workbook in memory
    wb_form = load_workbook(io.BytesIO(file_bytes), data_only=False)
    wb_form, _, _, _ = generate_excel_workbook(wb_form)
    out_buf = io.BytesIO()
    wb_form.save(out_buf)
    out_buf.seek(0)
    excel_base64 = base64.b64encode(out_buf.read()).decode("ascii")

    # Build daily_data & shortage_alerts for frontend compatibility
    daily_data = []
    for cabang, cap_val in wh_capacity:
        series_occ = occ_f.get(cabang, [0.0] * n_weeks)
        for w in range(n_weeks):
            occ_pct = series_occ[w] if series_occ[w] is not None else 0.0
            tot_bal = occ_pct * cap_val if cap_val else 0.0
            daily_data.append({
                "cabang": str(cabang),
                "date": period_labels[w],
                "total_on_hand": round(tot_bal, 2),
                "capacity": round(cap_val, 2),
                "occupancy_pct": round(occ_pct * 100, 2),
                "is_shortage": tot_bal < 0
            })

    shortage_alerts = []
    for rec in records:
        bf = bal_f[id(rec)]
        for w in range(n_weeks):
            if bf[w] < 0:
                shortage_alerts.append({
                    "cabang": rec.cabang,
                    "category": f"{rec.grup} - {rec.category}",
                    "date": period_labels[w],
                    "deficit": round(abs(bf[w]), 2)
                })

    avg_occ = sum(d["occupancy_pct"] for d in daily_data) / len(daily_data) if daily_data else 0
    max_occ = max(d["occupancy_pct"] for d in daily_data) if daily_data else 0

    # Build tidy dataframe for ABC-XYZ inventory engine
    inv_rows = []
    for rec in records:
        bf = bal_f[id(rec)]
        for w in range(n_weeks):
            # Using stable pseudo-dates to enable datetime sort in ABC-XYZ analysis
            synth_date = f"2026-01-{(w % 28) + 1:02d}"
            inv_rows.append({
                "Cabang": rec.cabang,
                "Category": f"{rec.grup} - {rec.category}",
                "Date": synth_date,
                "Penjualan": rec.forecast[w],
                "On Hand": bf[w],
                "PeriodLabel": period_labels[w]
            })

    import pandas as pd
    from services.inventory_engine import run_inventory_analysis
    inv_df = pd.DataFrame(inv_rows)
    inventory_result = run_inventory_analysis(inv_df) if not inv_df.empty else None

    return {
        "processed_at": datetime.now().isoformat(),
        "daily_data": daily_data,
        "branch_date_summary": daily_data,
        "shortage_alerts": shortage_alerts,
        "kpi_summary": {
            "avg_occupancy": round(avg_occ, 1),
            "max_occupancy": round(max_occ, 1),
            "categories_at_risk": len(shortage_alerts)
        },
        "over_occupancy_insights": insights,
        "inventory_analysis": inventory_result,
        "ddmrp_results": {
            "week_awal": week_awal,
            "period_labels": period_labels,
            "charts": {
                "balance_forecast": b64_bal_f,
                "balance_target": b64_bal_t,
                "occupancy_forecast": b64_occ_f,
                "occupancy_target": b64_occ_t,
            },
            "html_report": html_report,
            "excel_base64": excel_base64
        }
    }


# ============================================================================
# BAGIAN 4 - MAIN / CLI
# ============================================================================

def main():
    parser = argparse.ArgumentParser(description="DDMRP Program (gabungan Excel generator + analisa)")
    parser.add_argument("input", help="Path file Excel input (harus punya sheet Raw & WH)")
    parser.add_argument("--output-xlsx", default=None, help="Path file Excel output")
    parser.add_argument("--output-html", default=None, help="Path laporan HTML output")
    parser.add_argument("--charts-dir", default="charts", help="Folder untuk menyimpan grafik PNG")
    parser.add_argument("--mode", choices=["all", "excel", "analysis"], default="all",
                         help="'all' (default): generate Excel + laporan analisa. "
                              "'excel': hanya generate Excel. 'analysis': hanya laporan analisa.")
    args = parser.parse_args()

    base, ext = os.path.splitext(args.input)
    output_xlsx = args.output_xlsx or f"{base}_hasil{ext or '.xlsx'}"
    output_html = args.output_html or f"{base}_analysis_report.html"

    if args.mode in ("all", "excel"):
        generate_excel(args.input, output_xlsx)
    if args.mode in ("all", "analysis"):
        generate_analysis(args.input, output_html, args.charts_dir)


if __name__ == "__main__":
    main()
