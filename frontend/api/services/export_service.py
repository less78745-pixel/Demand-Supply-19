import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

def generate_excel_report(task_id: str) -> io.BytesIO:
    """
    Generates an Excel report using openpyxl.
    In a real scenario, task_id would be used to fetch results from a DB/Cache.
    Here we generate a mock structured report.
    """
    wb = Workbook()
    
    # Sheet 1: Summary
    ws1 = wb.active
    ws1.title = "Executive Summary"
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0B1120", end_color="0B1120", fill_type="solid")
    
    ws1['A1'] = "DSP Analysis Report"
    ws1['A1'].font = Font(bold=True, size=16)
    ws1.merge_cells('A1:C1')
    
    ws1['A3'] = "Task ID:"
    ws1['B3'] = task_id
    
    ws1['A5'] = "Metric"
    ws1['B5'] = "Value"
    for cell in ['A5', 'B5']:
        ws1[cell].font = header_font
        ws1[cell].fill = header_fill
        ws1[cell].alignment = Alignment(horizontal='center')
        
    ws1['A6'] = "Status"
    ws1['B6'] = "Completed Successfully"
    
    # Sheet 2: Data (Mock)
    ws2 = wb.create_sheet(title="Raw Data")
    
    headers = ["Date", "Category", "Value", "Notes"]
    for col_num, header in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        
    for row_num in range(2, 10):
        ws2.cell(row=row_num, column=1, value=f"2026-07-{row_num:02d}")
        ws2.cell(row=row_num, column=2, value="Category A")
        ws2.cell(row=row_num, column=3, value=row_num * 10)
        ws2.cell(row=row_num, column=4, value="Simulated data")
        
    # Adjust column widths
    for sheet in wb.worksheets:
        for column_cells in sheet.columns:
            length = max(len(str(cell.value)) for cell in column_cells if cell.value)
            sheet.column_dimensions[column_cells[0].column_letter].width = length + 2

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    
    return stream
