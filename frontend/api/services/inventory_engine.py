import pandas as pd
import numpy as np
import math
import io
from openpyxl import load_workbook

def _safe_float(v):
    """Safely cast to float and replace NaN or Inf with 0.0 to avoid JSON serialization errors."""
    try:
        val = float(v)
        if math.isnan(val) or math.isinf(val):
            return 0.0
        return val
    except Exception:
        return 0.0

def run_inventory_from_mrp_bytes(file_bytes: bytes) -> dict:
    """
    Ekstrak data mingguan dari sheet Raw & WH MRP untuk diolah ke dalam ABC-XYZ Inventory Intelligence.
    """
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    if "Raw" not in wb.sheetnames or "WH" not in wb.sheetnames:
        raise ValueError("File tidak memiliki sheet 'Raw' dan 'WH'")
    
    from services.occupancy_engine import detect_week_count, read_raw_records, compute_balance_series, read_week_awal, build_period_labels
    ws_raw = wb["Raw"]
    ws_wh = wb["WH"]
    n_weeks = detect_week_count(ws_raw)
    week_awal = read_week_awal(ws_wh, default=1)
    period_labels = build_period_labels(week_awal, n_weeks)
    records = read_raw_records(ws_raw, n_weeks)
    
    inv_rows = []
    for rec in records:
        bf = compute_balance_series(rec, n_weeks, "forecast")
        for w in range(n_weeks):
            synth_date = f"2026-01-{(w % 28) + 1:02d}"
            inv_rows.append({
                "Cabang": rec.cabang,
                "Category": f"{rec.grup} - {rec.category}",
                "Date": synth_date,
                "Penjualan": rec.forecast[w],
                "On Hand": bf[w],
                "PeriodLabel": period_labels[w]
            })
            
    df_inv = pd.DataFrame(inv_rows)
    return run_inventory_analysis(df_inv)

# Backward compatibility alias
run_inventory_from_ddmrp_bytes = run_inventory_from_mrp_bytes


def run_inventory_analysis(df: pd.DataFrame) -> dict:
    """
    ABC-XYZ Classification per Cabang × Category.
    Expects columns: Cabang, Category, Date, Penjualan, On Hand
    """
    if df.empty:
        return {"error": "Empty dataset"}

    df = df.copy()
    df['Date'] = pd.to_datetime(df['Date'], errors='coerce')

    if 'Cabang' not in df.columns:
        df['Cabang'] = 'Unknown'

    df['Cabang'] = df['Cabang'].astype(str)
    df['Category'] = df['Category'].astype(str)

    results = []
    cabangs = df['Cabang'].unique()

    for cabang in cabangs:
        cabang_df = df[df['Cabang'] == cabang]
        categories = cabang_df['Category'].unique()

        for cat in categories:
            cat_df = cabang_df[cabang_df['Category'] == cat].sort_values('Date')
            if cat_df.empty:
                continue

            sales        = pd.to_numeric(cat_df['Penjualan'], errors='coerce').fillna(0)
            total_volume = _safe_float(sales.sum())
            mean_sales   = _safe_float(sales.mean())
            std_sales    = _safe_float(sales.std(ddof=0))

            cv = std_sales / mean_sales if mean_sales > 0 else 0
            if cv <= 0.5:
                xyz = 'X'
            elif cv <= 1.0:
                xyz = 'Y'
            else:
                xyz = 'Z'

            on_hand_col = 'On Hand' if 'On Hand' in cat_df.columns else None
            current_on_hand = _safe_float(pd.to_numeric(cat_df[on_hand_col], errors='coerce').iloc[-1]) if on_hand_col else 0.0
            daily_sales = mean_sales / 30 if mean_sales > 0 else 0
            doh = current_on_hand / daily_sales if daily_sales > 0 else 9999

            stockout_risk = doh < 14
            overstock     = doh > 60
            
            if len(sales) >= 6:
                trend_pct = ((sales.iloc[-3:].mean() - sales.iloc[-6:-3].mean()) / (sales.iloc[-6:-3].mean() + 1e-9)) * 100
            elif len(sales) >= 2:
                trend_pct = ((sales.iloc[-1] - sales.iloc[0]) / (abs(sales.iloc[0]) + 1e-9)) * 100
            else:
                trend_pct = 0.0

            results.append({
                "cabang":         str(cabang),
                "category":       str(cat),
                "volume":         _safe_float(total_volume),
                "xyz_class":      xyz,
                "cv":             _safe_float(round(cv, 4)),
                "doh":            _safe_float(round(doh, 1)),
                "on_hand":        _safe_float(current_on_hand),
                "mean_sales":     _safe_float(round(mean_sales, 2)),
                "std_sales":      _safe_float(round(std_sales, 2)),
                "stockout_risk":  bool(stockout_risk),
                "overstock":      bool(overstock),
                "trend_pct":      _safe_float(round(trend_pct, 2)),
            })

    if not results:
        return {"error": "No valid data after processing"}

    final_results, dead_stock = [], []
    for cabang in cabangs:
        cab_res = [r for r in results if r['cabang'] == cabang]
        if not cab_res:
            continue

        df_cab = pd.DataFrame(cab_res).sort_values('volume', ascending=False).reset_index(drop=True)
        total_vol = df_cab['volume'].sum()

        if total_vol > 0:
            df_cab['cum_vol']  = df_cab['volume'].cumsum()
            df_cab['cum_perc'] = df_cab['cum_vol'] / total_vol
        else:
            df_cab['cum_perc'] = 0.0

        def get_abc(perc):
            if perc <= 0.80: return 'A'
            if perc <= 0.95: return 'B'
            return 'C'

        df_cab['abc_class'] = df_cab['cum_perc'].apply(get_abc)

        for _, row in df_cab.iterrows():
            abc_xyz = f"{row['abc_class']}{row['xyz_class']}"
            strategy = _recommend_strategy(row['abc_class'], row['xyz_class'], bool(row['stockout_risk']), bool(row['overstock']))
            entry = {
                "cabang":        str(row['cabang']),
                "category":      str(row['category']),
                "class":         abc_xyz,
                "abc":           str(row['abc_class']),
                "xyz":           str(row['xyz_class']),
                "volume":        _safe_float(row['volume']),
                "doh":           _safe_float(row['doh']),
                "on_hand":       _safe_float(row['on_hand']),
                "mean_sales":    _safe_float(row['mean_sales']),
                "cv":            _safe_float(row['cv']),
                "stockout_risk": bool(row['stockout_risk']),
                "overstock":     bool(row['overstock']),
                "trend_pct":     _safe_float(row['trend_pct']),
                "strategy":      strategy,
            }
            final_results.append(entry)

            if row['doh'] > 90:
                dead_stock.append({
                    "cabang":   str(row['cabang']),
                    "category": str(row['category']),
                    "doh":      _safe_float(row['doh']),
                    "on_hand":  _safe_float(row['on_hand']),
                    "class":    abc_xyz
                })

    a_count    = sum(1 for x in final_results if x['abc'] == 'A')
    z_count    = sum(1 for x in final_results if x['xyz'] == 'Z')
    risk_count = sum(1 for x in final_results if x['stockout_risk'])

    return {
        "matrix_data": final_results,
        "dead_stock":  dead_stock,
        "kpi_summary": {
            "total_categories": len(final_results),
            "a_class_count":    a_count,
            "z_class_count":    z_count,
            "dead_stock_count": len(dead_stock),
            "stockout_risk_count": risk_count,
        }
    }


def _recommend_strategy(abc: str, xyz: str, stockout: bool, overstock: bool) -> str:
    matrix = {
        ('A', 'X'): "Continuous review, tight safety stock. High value, predictable.",
        ('A', 'Y'): "Periodic review with buffer. Monitor lead-time closely.",
        ('A', 'Z'): "Collaborative forecasting needed. Risk of high-cost stockouts.",
        ('B', 'X'): "Standard replenishment. Regular cycle counts.",
        ('B', 'Y'): "Semi-annual review. Consider vendor-managed inventory.",
        ('B', 'Z'): "Review order frequency. Demand shaping may help.",
        ('C', 'X'): "Bulk ordering or min-max policy. Low priority.",
        ('C', 'Y'): "Reduce SKU complexity. Review necessity.",
        ('C', 'Z'): "Candidate for discontinuation or consignment stock.",
    }
    base = matrix.get((abc, xyz), "Standard inventory policy.")
    if stockout:
        base += " ⚠️ STOCKOUT RISK — replenish immediately."
    if overstock:
        base += " 📦 OVERSTOCK — consider promotion or redistribution."
    return base
