"""
DSP Calculator - Konversi Container -> QTY -> Rupiah
=====================================================

Modul ini menambahkan dua tahap konversi matematis di atas sheet
"1. Running Balance" (satuan container) milik menu Kalkulator DSP
(occupancy & inventory):

    Tahap 1 : Container -> QTY   (QTY = Nilai_Container * 68 / CBM)
    Tahap 2 : QTY       -> Rupiah (Rupiah = QTY * Harga Product)

Setiap tahap ditempatkan di sebelah kanan tabel sebelumnya, dipisahkan
oleh 10 kolom kosong (dummy) agar layout Excel tetap rapi dan mudah
dibaca oleh user.

CHAIN-OF-THOUGHT - Cara terbaik menyisipkan 10 kolom kosong di Excel
---------------------------------------------------------------------
Ada dua pendekatan yang mungkin:

1) Menyisipkan 10 kolom "dummy" berisi NaN ke dalam DataFrame lalu
   meng-export seluruh DataFrame sekaligus dengan `df.to_excel()`.
   Masalahnya:
     - Pandas mewajibkan nama kolom ada (walau kosong), sehingga baris
       header akan berisi label duplikat/aneh ("Unnamed: 0", dst).
     - NaN yang ditulis lewat `to_excel` tetap menghasilkan sel yang
       secara teknis "ada" (bisa perlu `na_rep=""`), bukan sel yang
       benar-benar kosong (Excel ISBLANK() bisa saja FALSE).

2) (DIPAKAI DI SCRIPT INI) Pisahkan proses "hitung" dari proses
   "tulis". Gunakan Pandas murni untuk kalkulasi (merge/lookup &
   operasi vektor), lalu tulis ke Excel dengan openpyxl secara
   per-blok menggunakan offset kolom (`start_col`, `start_col + gap`,
   dst). Kolom-kolom pada rentang gap TIDAK PERNAH disentuh sama
   sekali oleh proses penulisan.
     - Karena openpyxl hanya membuat Cell ketika kita menuliskan nilai
       ke sana, kolom yang dilewati (gap) otomatis menjadi sel kosong
       native (`None`), bukan string kosong - inilah definisi "benar-benar
       kosong" yang diminta.
     - Kita juga menghindari NaN tertulis sebagai string "nan": setiap
       nilai yang NaN sengaja di-skip saat penulisan sel (value tetap
       None) alih-alih dipaksa jadi 0 atau string kosong.

Pendekatan (2) inilah yang dipakai di seluruh fungsi `write_*` di bawah.
"""

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================================
# KONFIGURASI
# ============================================================================

FIXED_COLS = ["No", "Cabang", "Grup", "Category"]   # kolom identitas baris
N_FIXED_COLS = len(FIXED_COLS)
N_GAP_COLS = 10                                      # jumlah kolom kosong pemisah antar blok
FAKTOR_QTY = 68                                       # faktor pengali container -> qty

FILL_JUDUL = PatternFill("solid", fgColor="BFBFBF")
FILL_PERHITUNGAN = PatternFill("solid", fgColor="9DC3E6")
FILL_QTY = PatternFill("solid", fgColor="A9D18E")
FILL_RUPIAH = PatternFill("solid", fgColor="FFD966")
FILL_HEADER2 = PatternFill("solid", fgColor="D9D9D9")
BORDER = Border(*(Side(style="thin"),) * 4)
FONT_HEADER = Font(bold=True)


# ============================================================================
# 1. INISIALISASI RAW DATA SHEET (LOOKUP TABLES)
# ============================================================================

def build_lookup_cbm() -> pd.DataFrame:
    """Sheet 'CBM': volume (m3) per unit, per Cabang & Category."""
    data = [
        {"Cabang": "Jakarta",  "Category": "Elektronik", "Nilai": 0.85},
        {"Cabang": "Jakarta",  "Category": "Garmen",      "Nilai": 0.42},
        {"Cabang": "Surabaya", "Category": "Elektronik", "Nilai": 0.90},
        {"Cabang": "Surabaya", "Category": "Garmen",      "Nilai": 0.40},
        {"Cabang": "Medan",    "Category": "Elektronik", "Nilai": 0.88},
        # sengaja tidak ada baris "Medan - Garmen" untuk mensimulasikan lookup yang tidak match (NaN)
    ]
    return pd.DataFrame(data, columns=["Cabang", "Category", "Nilai"])


def build_lookup_harga_product() -> pd.DataFrame:
    """Sheet 'Harga Product': harga rupiah per qty, per Cabang & Category."""
    data = [
        {"Cabang": "Jakarta",  "Category": "Elektronik", "Nilai": 1_500_000},
        {"Cabang": "Jakarta",  "Category": "Garmen",      "Nilai": 250_000},
        {"Cabang": "Surabaya", "Category": "Elektronik", "Nilai": 1_450_000},
        {"Cabang": "Surabaya", "Category": "Garmen",      "Nilai": 240_000},
        {"Cabang": "Medan",    "Category": "Elektronik", "Nilai": 1_600_000},
        {"Cabang": "Medan",    "Category": "Garmen",      "Nilai": 260_000},
    ]
    return pd.DataFrame(data, columns=["Cabang", "Category", "Nilai"])


def build_sample_running_balance(week_cols) -> pd.DataFrame:
    """Contoh data sheet '1. Running Balance' dalam satuan container."""
    rows = [
        {"Cabang": "Jakarta",  "Grup": "FMCG",  "Category": "Elektronik"},
        {"Cabang": "Jakarta",  "Grup": "FMCG",  "Category": "Garmen"},
        {"Cabang": "Surabaya", "Grup": "Retail", "Category": "Elektronik"},
        {"Cabang": "Surabaya", "Grup": "Retail", "Category": "Garmen"},
        {"Cabang": "Medan",    "Grup": "Retail", "Category": "Elektronik"},
        {"Cabang": "Medan",    "Grup": "Retail", "Category": "Garmen"},   # -> tidak punya CBM lookup
    ]
    df = pd.DataFrame(rows)
    df.insert(0, "No", range(1, len(df) + 1))

    rng = np.random.default_rng(seed=42)
    for w in week_cols:
        df[w] = rng.integers(low=5, high=40, size=len(df))  # satuan container
    return df


# ============================================================================
# 2. KONVERSI TAHAP 1: CONTAINER -> QTY
# ============================================================================

def convert_container_to_qty(df_running_balance: pd.DataFrame, week_cols, df_cbm: pd.DataFrame) -> pd.DataFrame:
    """
    QTY = (Nilai Running Balance per week * 68) / Nilai CBM (lookup by Cabang & Category).

    Merge menggunakan how="left" agar seluruh baris Running Balance tetap
    dipertahankan. Baris yang tidak menemukan pasangan di sheet CBM akan
    menghasilkan CBM = NaN -> QTY juga NaN (bukan error), sehingga saat
    ditulis ke Excel selnya akan kosong, bukan 0 atau #DIV/0!.
    """
    merged = df_running_balance[["Cabang", "Category"]].merge(
        df_cbm.rename(columns={"Nilai": "CBM"}),
        on=["Cabang", "Category"],
        how="left",
    )

    # Hindari pembagian dengan 0 -> ganti 0 menjadi NaN supaya hasilnya NaN, bukan inf
    cbm_aman = merged["CBM"].replace(0, np.nan)

    qty_df = pd.DataFrame(index=df_running_balance.index)
    for w in week_cols:
        qty_df[w] = (df_running_balance[w] * FAKTOR_QTY) / cbm_aman

    return qty_df


# ============================================================================
# 3. KONVERSI TAHAP 2: QTY -> RUPIAH
# ============================================================================

def convert_qty_to_rupiah(df_qty: pd.DataFrame, week_cols, df_keys: pd.DataFrame, df_harga: pd.DataFrame) -> pd.DataFrame:
    """
    Rupiah = Nilai QTY per week * Nilai Harga Product (lookup by Cabang & Category).

    df_keys menyediakan kolom Cabang & Category (index selaras dengan df_qty)
    karena df_qty sendiri hanya berisi kolom-kolom week hasil konversi.
    """
    merged = df_keys[["Cabang", "Category"]].merge(
        df_harga.rename(columns={"Nilai": "Harga"}),
        on=["Cabang", "Category"],
        how="left",
    )
    harga = merged["Harga"]  # NaN otomatis jika tidak ada match

    rupiah_df = pd.DataFrame(index=df_qty.index)
    for w in week_cols:
        rupiah_df[w] = df_qty[w] * harga

    return rupiah_df


# ============================================================================
# 4. PENULISAN KE EXCEL (LAYOUTING DENGAN GAP 10 KOLOM)
# ============================================================================

def _style_cell(ws, row, col, value, fill=None, bold=False):
    c = ws.cell(row=row, column=col, value=value)
    c.border = BORDER
    c.alignment = Alignment(horizontal="center", vertical="center")
    if fill:
        c.fill = fill
    if bold:
        c.font = FONT_HEADER
    return c


def _write_group_title(ws, start_col, n_cols, title, fill):
    """Baris 1: judul grup (mis. 'Perhitungan (Running Balance)') di-merge sepanjang n_cols."""
    end_col = start_col + n_cols - 1
    ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
    _style_cell(ws, 1, start_col, title, fill=fill, bold=True)


def _write_block(ws, start_col, keys_df, value_df, week_cols, group_title, fill):
    """
    Menulis satu blok tabel (Perhitungan / QTY / Rupiah) mulai dari `start_col`.
    Kolom-kolom SEBELUM start_col dan SESUDAH kolom terakhir blok ini (yaitu
    area gap) tidak pernah disentuh oleh fungsi ini -> otomatis tetap kosong.
    """
    n_cols = len(week_cols)
    _write_group_title(ws, start_col, n_cols, group_title, fill)

    # Baris 2: header per-week
    for i, w in enumerate(week_cols):
        _style_cell(ws, 2, start_col + i, w, fill=FILL_HEADER2, bold=True)

    # Baris 3+: data. Nilai NaN sengaja TIDAK ditulis (cell.value tetap None)
    for r_idx, row_pos in enumerate(value_df.index):
        out_row = 3 + r_idx
        for i, w in enumerate(week_cols):
            val = value_df.at[row_pos, w]
            if pd.isna(val):
                _style_cell(ws, out_row, start_col + i, None)  # sel benar-benar kosong
            else:
                _style_cell(ws, out_row, start_col + i, round(float(val), 2))


def _write_fixed_identity_cols(ws, keys_df):
    """Menulis kolom identitas No/Cabang/Grup/Category di kolom 1..4."""
    _write_group_title(ws, 1, N_FIXED_COLS, "Judul", FILL_JUDUL)
    for i, title in enumerate(FIXED_COLS, start=1):
        _style_cell(ws, 2, i, title, fill=FILL_HEADER2, bold=True)

    for r_idx, (_, row) in enumerate(keys_df.iterrows()):
        out_row = 3 + r_idx
        for i, col_name in enumerate(FIXED_COLS, start=1):
            _style_cell(ws, out_row, i, row[col_name])


def export_running_balance_sheet(wb: Workbook, df_running_balance, week_cols, df_qty, df_rupiah):
    sheet_name = "1. Running Balance"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)

    keys_df = df_running_balance[FIXED_COLS]
    _write_fixed_identity_cols(ws, keys_df)

    # --- Blok 1: Perhitungan (Running Balance) - satuan container ---
    perhitungan_start_col = N_FIXED_COLS + 1
    _write_block(
        ws, perhitungan_start_col, keys_df, df_running_balance, week_cols,
        "Perhitungan (Running Balance)", FILL_PERHITUNGAN,
    )
    perhitungan_end_col = perhitungan_start_col + len(week_cols) - 1

    # --- Gap 10 kolom kosong, lalu Blok 2: QTY ---
    qty_start_col = perhitungan_end_col + 1 + N_GAP_COLS
    _write_block(ws, qty_start_col, keys_df, df_qty, week_cols, "QTY (Container -> QTY)", FILL_QTY)
    qty_end_col = qty_start_col + len(week_cols) - 1

    # --- Gap 10 kolom kosong lagi, lalu Blok 3: Rupiah ---
    rupiah_start_col = qty_end_col + 1 + N_GAP_COLS
    _write_block(ws, rupiah_start_col, keys_df, df_rupiah, week_cols, "Rupiah (QTY -> Rupiah)", FILL_RUPIAH)
    rupiah_end_col = rupiah_start_col + len(week_cols) - 1

    # Lebar kolom identitas & blok data; kolom gap sengaja DIBIARKAN default (tidak diset)
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 14
    for c in range(perhitungan_start_col, perhitungan_end_col + 1):
        ws.column_dimensions[get_column_letter(c)].width = 12
    for c in range(qty_start_col, qty_end_col + 1):
        ws.column_dimensions[get_column_letter(c)].width = 12
    for c in range(rupiah_start_col, rupiah_end_col + 1):
        ws.column_dimensions[get_column_letter(c)].width = 16

    ws.freeze_panes = "E3"
    return ws, {
        "perhitungan": (perhitungan_start_col, perhitungan_end_col),
        "qty": (qty_start_col, qty_end_col),
        "rupiah": (rupiah_start_col, rupiah_end_col),
    }


def export_lookup_sheet(wb: Workbook, sheet_name: str, df: pd.DataFrame):
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    for i, col in enumerate(df.columns, start=1):
        _style_cell(ws, 1, i, col, fill=FILL_HEADER2, bold=True)
    for r_idx, row in enumerate(df.itertuples(index=False), start=2):
        for i, val in enumerate(row, start=1):
            _style_cell(ws, r_idx, i, val)
    for i in range(1, len(df.columns) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 16
    return ws


# ============================================================================
# MAIN
# ============================================================================

def main():
    week_cols = [f"Week-{i}" for i in range(1, 5)]  # contoh 4 minggu

    df_cbm = build_lookup_cbm()
    df_harga_product = build_lookup_harga_product()
    df_running_balance = build_sample_running_balance(week_cols)

    # Tahap 1: Container -> QTY
    df_qty = convert_container_to_qty(df_running_balance, week_cols, df_cbm)

    # Tahap 2: QTY -> Rupiah
    df_rupiah = convert_qty_to_rupiah(df_qty, week_cols, df_running_balance[FIXED_COLS], df_harga_product)

    wb = Workbook()
    wb.remove(wb.active)  # buang sheet default kosong

    export_lookup_sheet(wb, "CBM", df_cbm)
    export_lookup_sheet(wb, "Harga Product", df_harga_product)
    ws, col_ranges = export_running_balance_sheet(wb, df_running_balance, week_cols, df_qty, df_rupiah)

    wb._sheets.sort(key=lambda s: {"CBM": 0, "Harga Product": 1, "1. Running Balance": 2}.get(s.title, 99))

    output_path = "dsp_calculator_running_balance.xlsx"
    wb.save(output_path)
    print(f"Berhasil menulis: {output_path}")
    print("Rentang kolom per blok (1-indexed):", col_ranges)


if __name__ == "__main__":
    main()
