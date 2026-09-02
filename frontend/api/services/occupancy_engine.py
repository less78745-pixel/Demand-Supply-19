import os
import io
import math
import base64
import numpy as np
from datetime import datetime, timedelta

from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.worksheet import Worksheet

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker

from sklearn.ensemble import IsolationForest
from .forecast_engine import _holt_winters_forecast

def _safe_float(v):
    try:
        val = float(v)
        if math.isnan(val) or math.isinf(val):
            return 0.0
        return val
    except Exception:
        return 0.0

# ============================================================================
# CORE MRP LOGIC (Sheet Raw & WH, Week Awal, Periode Dinamis)
# ============================================================================

FIXED_RAW_COLS = 4
ONHAND_COL = 5
FIRST_WEEK_BLOCK_COL = 6
COLS_PER_WEEK = 4

# Header row of the "Raw" sheet (see generate_mrp_template_bytes) -- used only
# to detect whether the newer template's "Region" column (inserted between
# No and Cabang) is present, so old uploads keep working untouched.
RAW_HEADER_ROW = 2

def _normalize_header(v) -> str:
    return str(v or "").strip().lower()

def resolve_raw_columns(ws_raw: Worksheet) -> dict:
    """Resolve 1-based column indices for the 'Raw' sheet's fixed columns
    (No, [Region], Cabang, Grup, Category, On Hand).

    This is a POSITION CHECK, not full header-driven parsing: it only peeks
    at column B of the header row (row 2) to see if it says "Region". If not,
    it returns the exact legacy layout (No,Cabang,Grup,Category,On Hand =
    columns 1-5) that every existing upload already relies on -- so files
    built from the old template are completely unaffected. Only when "Region"
    is actually found there does everything shift one column to the right."""
    header_row = list(next(
        ws_raw.iter_rows(min_row=RAW_HEADER_ROW, max_row=RAW_HEADER_ROW, values_only=True), []
    ))
    region_header = _normalize_header(header_row[1]) if len(header_row) > 1 else ""
    has_region = region_header in ("region", "regional", "area")
    if has_region:
        return {"no": 1, "region": 2, "cabang": 3, "grup": 4, "category": 5, "onhand": 6}
    return {"no": 1, "cabang": 2, "grup": 3, "category": 4, "onhand": 5}

MONTHS = ["JAN", "FEB", "MAR", "APR", "MAY", "JUN",
          "JUL", "AUG", "SEP", "OCT", "NOV", "DEC"]
WEEKS_PER_MONTH = 4

def period_label(week_number: int) -> str:
    from datetime import date
    if week_number < 1:
        week_number = 1
    
    # Calculate year and ISO week
    # Assuming base year 2024 (leap year, standard reference)
    year = 2024 + (week_number - 1) // 52
    iso_week = ((week_number - 1) % 52) + 1
    
    # Get the date of Monday for this ISO week
    try:
        d = date.fromisocalendar(year, iso_week, 1)
    except Exception:
        # Fallback if fromisocalendar is not available or errors out
        d = date(year, 1, 1)
        
    month_name = MONTHS[d.month - 1]
    
    # Calculate week of the month (1 to 5)
    week_in_month = (d.day - 1) // 7 + 1
    
    label = f"{month_name}-{week_in_month}"
    year_offset = year - 2024
    if year_offset > 0:
        label += f" (Y+{year_offset})"
    return label

def read_week_awal(ws_wh: Worksheet, default: int = 1) -> int:
    for row in ws_wh.iter_rows(values_only=True):
        for i, val in enumerate(row):
            if isinstance(val, str) and val.strip().lower() == "week awal":
                if i + 1 < len(row) and isinstance(row[i + 1], (int, float)):
                    return int(row[i + 1])
    return default

def detect_week_count(ws_raw: Worksheet, cols: dict = None) -> int:
    cols = cols or resolve_raw_columns(ws_raw)
    first_week_col = cols["onhand"] + 1
    max_col = ws_raw.max_column
    remaining = max_col - (first_week_col - 1)
    n_weeks = remaining // COLS_PER_WEEK
    return max(1, n_weeks)

def raw_week_cols(w: int, first_week_col: int = FIRST_WEEK_BLOCK_COL):
    base = first_week_col + COLS_PER_WEEK * w
    return base, base + 1, base + 2, base + 3

def get_raw_rows(ws_raw: Worksheet, cols: dict = None):
    cols = cols or resolve_raw_columns(ws_raw)
    no_col, cabang_col = cols["no"], cols["cabang"]
    max_col = max(no_col, cabang_col)
    rows = []
    for idx, row in enumerate(ws_raw.iter_rows(min_row=3, max_col=max_col, values_only=True), start=3):
        val1 = row[no_col - 1] if len(row) >= no_col else None
        val2 = row[cabang_col - 1] if len(row) >= cabang_col else None
        if val1 in (None, "") and val2 in (None, ""):
            continue
        if str(val2).strip().lower() in ("cabang", "branch", "grup"):
            continue
        rows.append(idx)
    return rows

def get_wh_rows(ws_wh: Worksheet):
    rows = []
    for idx, row in enumerate(ws_wh.iter_rows(min_row=2, max_col=2, values_only=True), start=2):
        val1 = row[0] if len(row) > 0 else None
        cabang = row[1] if len(row) > 1 else None
        if cabang in (None, ""):
            continue
        rows.append(idx)
    return rows

class RawRecord:
    __slots__ = ("no", "region", "cabang", "grup", "category", "onhand", "to", "vessel", "target")

    def __init__(self, no, cabang, grup, category, onhand, to, vessel, target, region=None):
        self.no = no
        # Absent on files built from the pre-Region template (see
        # resolve_raw_columns) -- defaults to "Unknown" rather than raising,
        # same fallback already used below for cabang/grup/category.
        self.region = str(region).strip() if region not in (None, "") else "Unknown"
        # Stripped here (not just at lookup time like the CBM/Harga Product
        # keys below) because cabang/grup/category are ALSO used directly as
        # dict/group keys throughout this module (qty_series_by_branch,
        # agg_bal_t, breakdown_by_cabang_week, ...). A stray trailing/leading
        # space in the "Raw" sheet's Cabang cell (e.g. "Bengkulu ") used to
        # silently split that branch into its own mismatched key - it would
        # get computed, but then fail to match the trimmed cabang name coming
        # from the "WH" sheet/filter list, so it disappeared specifically
        # from cabang-keyed views like MOS/Value while still showing up in
        # views driven by the WH capacity list.
        self.cabang = str(cabang).strip() if cabang is not None else "Unknown"
        self.grup = str(grup).strip() if grup is not None else "General"
        self.category = str(category).strip() if category is not None else "General"
        self.onhand = _safe_float(onhand)
        self.to = [_safe_float(x) for x in to]
        self.vessel = [_safe_float(x) for x in vessel]
        self.target = [_safe_float(x) for x in target]

    @property
    def label(self):
        return f"{self.cabang}-{self.grup}-{self.category}"

def read_raw_records(ws_raw: Worksheet, n_weeks: int, cols: dict = None):
    cols = cols or resolve_raw_columns(ws_raw)
    no_col = cols["no"]
    region_col = cols.get("region")
    cabang_col, grup_col, cat_col, onhand_col = cols["cabang"], cols["grup"], cols["category"], cols["onhand"]
    first_week_col = onhand_col + 1

    def col_val(row_vals, col_idx_1_based):
        idx = col_idx_1_based - 1
        return row_vals[idx] if idx < len(row_vals) else None

    records = []
    for row_vals in ws_raw.iter_rows(min_row=3, values_only=True):
        val_no = col_val(row_vals, no_col)
        val_cabang = col_val(row_vals, cabang_col)
        if val_no in (None, "") and val_cabang in (None, ""):
            continue
        if str(val_cabang).strip().lower() in ("cabang", "branch", "grup"):
            continue

        def get_val(col_idx_1_based):
            v = col_val(row_vals, col_idx_1_based)
            return v if v is not None else 0

        onhand = get_val(onhand_col)
        to, vessel, target = [], [], []
        for w in range(n_weeks):
            c_to, c_vessel, c_buffer, c_target = raw_week_cols(w, first_week_col)
            to.append(get_val(c_to))
            vessel.append(get_val(c_vessel))
            target.append(get_val(c_target))

        records.append(RawRecord(
            no=val_no,
            region=col_val(row_vals, region_col) if region_col else None,
            cabang=val_cabang,
            grup=col_val(row_vals, grup_col),
            category=col_val(row_vals, cat_col),
            onhand=onhand, to=to, vessel=vessel, target=target,
        ))
    return records

def read_wh_capacity(ws_wh: Worksheet):
    result = []
    for row in ws_wh.iter_rows(min_row=2, max_col=4, values_only=True):
        val = row[0] if len(row) > 0 else None
        cabang = row[1] if len(row) > 1 else None
        if cabang in (None, ""):
            continue
        existing = row[2] if len(row) > 2 and row[2] is not None else 0
        tambahan = row[3] if len(row) > 3 and row[3] is not None else 0
        try:
            total_cap = float(existing) + float(tambahan)
        except (ValueError, TypeError):
            total_cap = 0.0
        result.append((str(cabang).strip(), total_cap))
    return result

def compute_balance_series(record: RawRecord, n_weeks: int):
    demand = record.target
    balances = []
    prev = 0.0
    for w in range(n_weeks):
        if w == 0:
            bal = record.onhand + record.to[0] + record.vessel[0] - demand[0]
        else:
            bal = prev + record.to[w] + record.vessel[w] - demand[w]
        balances.append(bal)
        prev = bal
    return balances

def compute_ratio_series(balances, record: RawRecord, n_weeks: int):
    demand = record.target
    ratios = []
    for w in range(n_weeks):
        if w + 1 < n_weeks and demand[w + 1]:
            ratios.append(balances[w] / demand[w + 1])
        else:
            ratios.append(None)
    return ratios

def compute_occupancy(records, balances_by_record, wh_capacity):
    if not balances_by_record:
        return {}
    n_weeks = len(next(iter(balances_by_record.values())))
    occupancy = {}
    for cabang, capacity in wh_capacity:
        totals = [0.0] * n_weeks
        for rec in records:
            if rec.cabang == cabang:
                bals = balances_by_record[id(rec)]
                for w in range(n_weeks):
                    # Akurasi: shortage/minus tidak membebaskan ruang fisik gudang
                    totals[w] += max(0.0, bals[w])
        occupancy[cabang] = [t / capacity if capacity else None for t in totals]
    return occupancy

def build_period_labels(week_awal: int, n_weeks: int):
    return [period_label(week_awal + w) for w in range(n_weeks)]

# ============================================================================
# EXCEL SHEET BUILDERS & CHARTS & HTML REPORT
# ============================================================================

FILL_JUDUL = PatternFill("solid", fgColor="D9D9D9")
FILL_PERHITUNGAN = PatternFill("solid", fgColor="FCE4D6")
FILL_RATIO = PatternFill("solid", fgColor="CFE2F3")
FILL_HEADER2 = PatternFill("solid", fgColor="F2F2F2")
FILL_QTY = PatternFill("solid", fgColor="A9D18E")
FILL_RUPIAH = PatternFill("solid", fgColor="FFD966")

# Faktor konversi Container -> QTY, sama seperti dsp_calculator_conversion.py
FAKTOR_QTY = 68

def compute_qty_rupiah_series(cabang, category, bal_series, n_weeks, cbm_dict, harga_product_dict):
    """QTY = Running Balance (container) * FAKTOR_QTY / CBM (lookup Cabang+Category).
    Rupiah = QTY * Harga Product (lookup Cabang+Category).
    None saat CBM/Harga tidak ditemukan (lookup key tidak ada) atau CBM=0 --
    dibiarkan blank di Excel/JSON, bukan dipaksa 0, supaya "tidak ada data harga"
    tetap bisa dibedakan dari "nilainya nol"."""
    key = (str(cabang).strip(), str(category).strip())
    cbm_val = cbm_dict.get(key)
    harga_val = harga_product_dict.get(key)
    qty_series, rupiah_series = [], []
    for w in range(n_weeks):
        qty_val = (bal_series[w] * FAKTOR_QTY / cbm_val) if cbm_val else None
        rupiah_val = (qty_val * harga_val) if (qty_val is not None and harga_val is not None) else None
        qty_series.append(qty_val)
        rupiah_series.append(rupiah_val)
    return qty_series, rupiah_series

def compute_mos_value_series_by_branch(records, n_weeks, cbm_dict, harga_product_dict, value_series_by_branch):
    """MOS berbasis Value untuk metrik 'Analisa Nilai Inventori' (Fitur 3):

    1. Variabel AA (per record, per minggu) = Target QTY * Harga Product * 4.
       Target QTY didapat dari Target mentah (sheet Raw, satuan Container)
       dikonversi ke satuan QTY dengan rumus & lookup CBM (Cabang+Category)
       yang SAMA PERSIS dengan compute_qty_rupiah_series (Target * FAKTOR_QTY
       / CBM) -- WAJIB disamakan karena Value(Rp) di pembilang (dari
       value_series_by_branch) juga dihitung dari Balance yang sudah
       dikonversi ke QTY, bukan dari Balance mentah dalam Container. Tanpa
       konversi ini AA berada di satuan Container (jauh lebih kecil dari QTY,
       biasanya puluhan-ribu kali lipat tergantung CBM), sehingga SUM(AA)
       terlalu kecil dan MOS = Value/SUM(AA) meledak jadi jauh lebih besar
       dari yang seharusnya -- persis gejala yang dilaporkan ("hasilnya besar
       sekali"). Faktor `* 4` dikalikan DI SINI, per record, SEBELUM dijumlah
       -- bukan diterapkan ke SUM(AA) sesudahnya (hasil akhirnya sama karena
       4 konstanta, tapi urutan ini yang diminta).
    2. AA dijumlahkan (SUM) per Cabang per minggu, lintas semua Category/Grup di
       cabang tsb -- SUM(AA) merepresentasikan total nilai rupiah dari demand
       mingguan cabang itu (sudah termasuk faktor `* 4` dari langkah 1).
    3. MOS (Value) per Cabang per minggu = Value(Rp) dari `value_series_by_branch`
       (hasil modul Analisa Nilai Inventori) / SUM(AA), TANPA dikali 4 lagi di
       langkah ini (sudah dikali di langkah 1). Fallback ke 0 saat SUM(AA) = 0,
       supaya pembagian oleh nol tidak melempar exception.
    """
    aa_sum_by_branch = {}
    for rec in records:
        key = (str(rec.cabang).strip(), str(rec.category).strip())
        cbm_val = cbm_dict.get(key)
        harga = harga_product_dict.get(key, 0.0)
        acc = aa_sum_by_branch.setdefault(rec.cabang, [0.0] * n_weeks)
        for w in range(n_weeks):
            target_qty = (rec.target[w] * FAKTOR_QTY / cbm_val) if cbm_val else 0.0
            acc[w] += target_qty * harga * 4

    mos_value_series_by_branch = {}
    for cabang, value_series in value_series_by_branch.items():
        aa_sum = aa_sum_by_branch.get(cabang, [0.0] * n_weeks)
        series = []
        for w in range(n_weeks):
            denom = aa_sum[w]
            series.append(round(value_series[w] / denom, 4) if denom != 0 else 0.0)
        mos_value_series_by_branch[cabang] = series
    return mos_value_series_by_branch

NATIONAL_KEY = "NASIONAL"

def aggregate_qty_value_by_group(records, n_weeks, qty_rupiah_by_record, group_key_fn):
    """Generalisasi dari loop qty_series_by_branch/value_series_by_branch di
    calculate_mrp_occupancy_from_bytes: agregasi QTY & Value per kelompok
    arbitrary (per cabang, per region, atau satu key NATIONAL_KEY untuk total
    nasional), reuse qty_series/rupiah_series yang SUDAH dihitung per-record
    (qty_rupiah_by_record) supaya rumus konversi Container->QTY->Value persis
    sama di semua level agregasi -- tidak dihitung ulang dengan cara lain."""
    qty_by_group, value_by_group = {}, {}
    for rec in records:
        key = group_key_fn(rec)
        qty_series, rupiah_series = qty_rupiah_by_record[id(rec)]
        qty_acc = qty_by_group.setdefault(key, [0.0] * n_weeks)
        value_acc = value_by_group.setdefault(key, [0.0] * n_weeks)
        for w in range(n_weeks):
            if qty_series[w] is not None:
                qty_acc[w] += qty_series[w]
            if rupiah_series[w] is not None:
                value_acc[w] += rupiah_series[w]
    return qty_by_group, value_by_group

def compute_mos_value_series_by_group(records, n_weeks, cbm_dict, harga_product_dict,
                                        value_series_by_group, group_key_fn):
    """Generalisasi compute_mos_value_series_by_branch di atas -- logika AA
    (Target QTY x Harga Product x 4) & pembagian identik, group key-nya
    parametrized supaya bisa dipakai untuk level Cabang, Region, atau
    Nasional dengan SATU implementasi (lihat docstring fungsi di atas untuk
    penjelasan lengkap langkah 1-3)."""
    aa_sum_by_group = {}
    for rec in records:
        key = group_key_fn(rec)
        lookup_key = (str(rec.cabang).strip(), str(rec.category).strip())
        cbm_val = cbm_dict.get(lookup_key)
        harga = harga_product_dict.get(lookup_key, 0.0)
        acc = aa_sum_by_group.setdefault(key, [0.0] * n_weeks)
        for w in range(n_weeks):
            target_qty = (rec.target[w] * FAKTOR_QTY / cbm_val) if cbm_val else 0.0
            acc[w] += target_qty * harga * 4

    mos_by_group = {}
    for key, value_series in value_series_by_group.items():
        aa_sum = aa_sum_by_group.get(key, [0.0] * n_weeks)
        mos_by_group[key] = [
            round(value_series[w] / aa_sum[w], 4) if aa_sum[w] != 0 else 0.0
            for w in range(n_weeks)
        ]
    return mos_by_group

BOLD = Font(bold=True)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def style_header_cell(ws, row, col, text, fill=None):
    c = ws.cell(row=row, column=col, value=text)
    c.font = BOLD
    c.alignment = CENTER
    c.border = BORDER
    if fill:
        c.fill = fill
    return c

def build_step1_sheet(wb, ws_raw, raw_rows, n_weeks, period_labels, cols: dict = None):
    cols = cols or resolve_raw_columns(ws_raw)
    # Output sheet layout (No, Cabang, Grup, Category) is unchanged regardless
    # of whether the source "Raw" sheet has a Region column -- only the source
    # cell references below need to follow the resolved columns.
    sheet_name = "1. Running Balance"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    col_perhitungan_start = FIXED_RAW_COLS + 1
    total_cols = col_perhitungan_start + n_weeks - 1

    for c in range(1, FIXED_RAW_COLS + 1):
        style_header_cell(ws, 1, c, "Judul", FILL_JUDUL)
    for w in range(n_weeks):
        style_header_cell(ws, 1, col_perhitungan_start + w, "Perhitungan (Running Balance)", FILL_PERHITUNGAN)

    fixed_titles = ["No", "Cabang", "Grup", "Category"]
    for i, t in enumerate(fixed_titles, start=1):
        style_header_cell(ws, 2, i, t, FILL_HEADER2)
    for w in range(n_weeks):
        style_header_cell(ws, 2, col_perhitungan_start + w, period_labels[w], FILL_HEADER2)

    L_no = get_column_letter(cols["no"])
    L_cabang = get_column_letter(cols["cabang"])
    L_grup = get_column_letter(cols["grup"])
    L_category = get_column_letter(cols["category"])
    L_onhand = get_column_letter(cols["onhand"])
    first_week_col = cols["onhand"] + 1

    perhitungan_col_letters = [get_column_letter(col_perhitungan_start + w) for w in range(n_weeks)]
    raw_col_letters = []
    for w in range(n_weeks):
        col_to, col_vessel, col_buffer, col_target = raw_week_cols(w, first_week_col)
        raw_col_letters.append({
            "to": get_column_letter(col_to),
            "vessel": get_column_letter(col_vessel),
            "demand": get_column_letter(col_target)
        })

    out_row = 3
    for raw_row in raw_rows:
        row_vals = [f"=Raw!{L_no}{raw_row}", f"=Raw!{L_cabang}{raw_row}", f"=Raw!{L_grup}{raw_row}", f"=Raw!{L_category}{raw_row}"]
        for w in range(n_weeks):
            L_to = raw_col_letters[w]["to"]
            L_vessel = raw_col_letters[w]["vessel"]
            L_demand = raw_col_letters[w]["demand"]
            if w == 0:
                formula = f"=SUM(Raw!{L_onhand}{raw_row},Raw!{L_to}{raw_row},Raw!{L_vessel}{raw_row})-Raw!{L_demand}{raw_row}"
            else:
                prev_letter = perhitungan_col_letters[w - 1]
                formula = f"=SUM({prev_letter}{out_row},Raw!{L_to}{raw_row},Raw!{L_vessel}{raw_row})-Raw!{L_demand}{raw_row}"
            row_vals.append(formula)

        ws.append(row_vals)
        out_row += 1

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 16
    for c in range(col_perhitungan_start, total_cols + 1):
        ws.column_dimensions[get_column_letter(c)].width = 12

    ws.freeze_panes = "E3"
    ws.views.sheetView[0].showGridLines = True
    return ws, col_perhitungan_start, out_row - 1

def append_qty_rupiah_blocks(ws, raw_rows, n_weeks, period_labels, perhitungan_col_start, records, qty_rupiah_by_record):
    """Tambahkan blok 'QTY (Container -> QTY)' dan 'Rupiah (QTY -> Rupiah)' di
    sebelah kanan blok Perhitungan (Running Balance) pada sheet
    '1. Running Balance', mengikuti logika konversi dsp_calculator_conversion.py.

    Ditulis sebagai NILAI STATIS (bukan formula Excel live seperti blok
    Perhitungan) dengan sengaja -- sheet lookup CBM/Harga Product bisa berisi
    puluhan ribu baris di file nyata (satu upload user tercatat 51.275 &
    49.621 baris). SUMIFS/VLOOKUP whole-column terhadap tabel selebar itu untuk
    setiap sel per-minggu per-baris Raw akan membuat Excel sangat lambat/hang
    saat recalculate; lookup dict Python menyelesaikan hal yang sama sekali
    saat generate, jauh lebih murah."""
    if not raw_rows:
        return
    n_cols = n_weeks
    gap = 10
    perhitungan_end_col = perhitungan_col_start + n_weeks - 1
    qty_start_col = perhitungan_end_col + 1 + gap
    rupiah_start_col = qty_start_col + n_cols + gap

    def write_group_title(start_col, title, fill):
        end_col = start_col + n_cols - 1
        ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
        c = ws.cell(row=1, column=start_col, value=title)
        c.font = BOLD
        c.alignment = CENTER
        c.border = BORDER
        c.fill = fill

    write_group_title(qty_start_col, "QTY (Container -> QTY)", FILL_QTY)
    write_group_title(rupiah_start_col, "Rupiah (QTY -> Rupiah)", FILL_RUPIAH)

    for w in range(n_weeks):
        style_header_cell(ws, 2, qty_start_col + w, period_labels[w], FILL_HEADER2)
        style_header_cell(ws, 2, rupiah_start_col + w, period_labels[w], FILL_HEADER2)

    # Tulis nilai TANPA border/number_format per-sel: untuk upload nasional
    # (puluhan ribu `records` x banyak minggu) mengeset .border/.number_format
    # pada tiap sel satu-per-satu adalah bottleneck openpyxl yang nyata --
    # tiap assignment style memicu lookup/rebuild style-array internal, jadi
    # untuk records x n_weeks x 2 sel biayanya berlipat ganda dengan cepat.
    # Gridlines sheet tetap ON (lihat showGridLines di bawah) jadi kehilangan
    # border tipis per-sel tidak mengorbankan keterbacaan. number_format untuk
    # kolom Rupiah cukup diset SEKALI di level kolom (bukan per-sel) --
    # openpyxl/Excel memakainya sebagai default untuk sel tanpa format eksplisit.
    for out_row, rec in enumerate(records, start=3):
        qty_series, rupiah_series = qty_rupiah_by_record.get(id(rec), ([None] * n_weeks, [None] * n_weeks))
        for w in range(n_weeks):
            qty_val = qty_series[w]
            rupiah_val = rupiah_series[w]
            ws.cell(row=out_row, column=qty_start_col + w, value=round(qty_val, 2) if qty_val is not None else None)
            ws.cell(row=out_row, column=rupiah_start_col + w, value=round(rupiah_val, 2) if rupiah_val is not None else None)

    for c in range(qty_start_col, qty_start_col + n_weeks):
        ws.column_dimensions[get_column_letter(c)].width = 12
    for c in range(rupiah_start_col, rupiah_start_col + n_weeks):
        col_letter = get_column_letter(c)
        ws.column_dimensions[col_letter].width = 16
        ws.column_dimensions[col_letter].number_format = "#,##0"

def build_step2_sheet(wb, ws_wh, wh_rows, n_weeks, period_labels, perhitungan_col_start, hasil_last_row):
    sheet_name = "2. Occupancy"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    titles = ["No", "Cabang"] + period_labels
    for i, t in enumerate(titles, start=1):
        style_header_cell(ws, 1, i, t, FILL_HEADER2)
    out_row = 2
    perhitungan_col_letters = [get_column_letter(perhitungan_col_start + w) for w in range(n_weeks)]
    for wh_row in wh_rows:
        row_vals = [f"=WH!A{wh_row}", f"=WH!B{wh_row}"]
        for w in range(n_weeks):
            col_letter = perhitungan_col_letters[w]
            # SUMPRODUCT (bukan SUMIF) supaya balance negatif (shortage) di-floor ke 0
            # sebelum dijumlahkan -- meniru persis `max(0.0, bals[w])` di compute_occupancy()
            # (baris ~206) sehingga sheet ini konsisten dengan angka dashboard/API.
            rng_cabang = f"'1. Running Balance'!$B$3:$B${hasil_last_row}"
            rng_val = f"'1. Running Balance'!${col_letter}$3:${col_letter}${hasil_last_row}"
            formula = (
                f"=IFERROR(IF(WH!$E{wh_row}=0,0,"
                f"SUMPRODUCT(({rng_cabang}=$B{out_row})*({rng_val}>0)*{rng_val})"
                f"/WH!$E{wh_row}),0)"
            )
            row_vals.append(formula)
        ws.append(row_vals)
        for w in range(n_weeks):
            ws.cell(row=out_row, column=3 + w).number_format = "0.0%"
        out_row += 1

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 16
    for w in range(n_weeks):
        ws.column_dimensions[get_column_letter(3 + w)].width = 13

    ws.freeze_panes = "C2"
    ws.views.sheetView[0].showGridLines = True
    return ws

def build_step3_sheet(wb, ws_raw, raw_rows, n_weeks, period_labels, perhitungan_col_start, cols: dict = None):
    cols = cols or resolve_raw_columns(ws_raw)
    sheet_name = "3. Harga & MOS"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    headers = ["No", "Cabang", "Grup", "Week", "Balance", "Target", "Harga", "Value_per_Week", "MOS"]
    for i, h in enumerate(headers, start=1):
        style_header_cell(ws, 1, i, h, FILL_HEADER2)

    unique_groups = []
    seen = set()
    for r in raw_rows:
        cab = ws_raw.cell(row=r, column=cols["cabang"]).value
        grp = ws_raw.cell(row=r, column=cols["grup"]).value
        key = (str(cab).strip() if cab else "", str(grp).strip() if grp else "")
        if key not in seen and key[0]:
            seen.add(key)
            unique_groups.append(key)

    out_row = 2
    perhitungan_col_letters = [get_column_letter(perhitungan_col_start + w) for w in range(n_weeks)]
    first_week_col = cols["onhand"] + 1
    L_raw_cabang = get_column_letter(cols["cabang"])
    L_raw_grup = get_column_letter(cols["grup"])
    raw_col_letters = []
    for w in range(n_weeks):
        _, _, _, col_target = raw_week_cols(w, first_week_col)
        raw_col_letters.append(get_column_letter(col_target))

    idx = 1
    for cab, grp in unique_groups:
        for w in range(n_weeks):
            week_label = period_labels[w]
            ht_col = perhitungan_col_letters[w]
            raw_col = raw_col_letters[w]

            f_balance = f"=SUMIFS('1. Running Balance'!${ht_col}:${ht_col}, '1. Running Balance'!$B:$B, B{out_row}, '1. Running Balance'!$C:$C, C{out_row})"
            f_target = f"=SUMIFS('Raw'!${raw_col}:${raw_col}, 'Raw'!${L_raw_cabang}:${L_raw_cabang}, B{out_row}, 'Raw'!${L_raw_grup}:${L_raw_grup}, C{out_row})"
            f_harga = f"=SUMIFS('Harga Container'!$D:$D, 'Harga Container'!$B:$B, B{out_row}, 'Harga Container'!$C:$C, C{out_row})"
            f_val = f"=E{out_row} * G{out_row}"
            # MOS = Balance / (Target * 4) -- lihat mos_rows di
            # calculate_mrp_occupancy_from_bytes untuk penjelasan lengkap.
            f_mos = f'=IFERROR(E{out_row} / (F{out_row} * 4), 0)'
            
            ws.append([idx, cab, grp, week_label, f_balance, f_target, f_harga, f_val, f_mos])
            
            idx += 1
            out_row += 1

    for c in range(1, 10):
        ws.column_dimensions[get_column_letter(c)].width = 15
    ws.freeze_panes = "A2"
    ws.views.sheetView[0].showGridLines = True
    return ws

def generate_excel_workbook(wb: Workbook, records=None, qty_rupiah_by_record=None):
    ws_raw, ws_wh = wb["Raw"], wb["WH"]
    cols = resolve_raw_columns(ws_raw)
    n_weeks = detect_week_count(ws_raw, cols)
    raw_rows, wh_rows = get_raw_rows(ws_raw, cols), get_wh_rows(ws_wh)
    week_awal = read_week_awal(ws_wh, default=1)
    period_labels = build_period_labels(week_awal, n_weeks)

    for r in wh_rows:
        ws_wh.cell(row=r, column=5, value=f"=C{r}+D{r}")

    ws_step1, ht_start, ht_last = build_step1_sheet(wb, ws_raw, raw_rows, n_weeks, period_labels, cols)

    if records is not None and qty_rupiah_by_record is not None and len(records) == len(raw_rows):
        append_qty_rupiah_blocks(ws_step1, raw_rows, n_weeks, period_labels, ht_start, records, qty_rupiah_by_record)

    build_step2_sheet(wb, ws_wh, wh_rows, n_weeks, period_labels, ht_start, ht_last)
    build_step3_sheet(wb, ws_raw, raw_rows, n_weeks, period_labels, ht_start, cols)

    order = ["Raw", "WH", "Harga Container", "CBM", "Harga Product", "1. Running Balance", "2. Occupancy", "3. Harga & MOS"]
    wb._sheets.sort(key=lambda s: order.index(s.title) if s.title in order else 999)
    return wb, period_labels, week_awal, n_weeks

CATEGORICAL = ["#2a78d6", "#eb6834", "#1baf7a", "#eda100", "#e87ba4", "#008300", "#4a3aa7", "#e34948"]
SURFACE, INK_PRIMARY, INK_SECONDARY, INK_MUTED, GRIDLINE, BASELINE, STATUS_CRITICAL = "#fcfcfb", "#0b0b0b", "#52514e", "#898781", "#e1e0d9", "#c3c2b7", "#d03b3b"

def _style_axes(ax):
    ax.grid(axis="y", which="major")
    ax.grid(axis="x", visible=False)
    ax.spines["left"].set_color(BASELINE)
    ax.spines["bottom"].set_color(BASELINE)

def plot_balance_chart_b64(records, balances_by_record, period_labels, title):
    fig, ax = plt.subplots(figsize=(9.5, 5))
    _style_axes(ax)
    
    # A.1: Group by Cabang & Grup
    agg_bals = {}
    for rec in records:
        group_key = f"{rec.cabang}-{rec.grup}"
        if group_key not in agg_bals:
            agg_bals[group_key] = [0.0] * len(period_labels)
        bals = balances_by_record[id(rec)]
        for w in range(len(period_labels)):
            agg_bals[group_key][w] += bals[w]
            
    plot_keys = list(agg_bals.keys())[:15] if len(agg_bals) > 15 else list(agg_bals.keys())
    for i, key in enumerate(plot_keys):
        ax.plot(period_labels, agg_bals[key], marker="o", markersize=5, linewidth=2, color=CATEGORICAL[i % len(CATEGORICAL)], label=key)
    ax.axhline(0, color=STATUS_CRITICAL, linewidth=1, linestyle="--", alpha=0.6)
    ax.set_ylabel("Balance (unit)")
    ax.set_title(title, fontsize=13, fontweight="bold", color=INK_PRIMARY, loc="left")
    ax.legend(loc="upper left", bbox_to_anchor=(1.0, 1.0), frameon=False, fontsize=9)
    fig.subplots_adjust(right=0.78)
    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=150, facecolor=SURFACE)
    plt.close(fig)
    buf.seek(0)
    return base64.b64encode(buf.read()).decode("ascii")

def plot_occupancy_chart_b64(occupancy, period_labels, title):
    fig, ax = plt.subplots(figsize=(8.5, 5))
    _style_axes(ax)
    for i, cabang in enumerate(occupancy.keys()):
        y = [v * 100 if v is not None else None for v in occupancy[cabang]]
        ax.plot(period_labels, y, marker="o", markersize=6, linewidth=2.2, color=CATEGORICAL[i % len(CATEGORICAL)], label=cabang)
    ax.axhline(100, color=STATUS_CRITICAL, linewidth=1.4, linestyle="--")
    ax.text(0.01, 100, " Kapasitas 100% ", color=STATUS_CRITICAL, fontsize=9, va="bottom", ha="left", transform=ax.get_yaxis_transform())
    ax.yaxis.set_major_formatter(mticker.PercentFormatter())
    ax.set_ylabel("Occupancy (%)")
    ax.set_title(title, fontsize=13, fontweight="bold", color=INK_PRIMARY, loc="left")
    ax.legend(loc="upper left", bbox_to_anchor=(1.0, 1.0), frameon=False, fontsize=9)
    fig.subplots_adjust(right=0.8)
    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=150, facecolor=SURFACE)
    plt.close(fig)
    buf.seek(0)
    return base64.b64encode(buf.read()).decode("ascii")

def generate_insights(records, bal_t, ratio_t, occ_t, period_labels, n_weeks):
    insights = []
    for rec in records:
        for w, r in enumerate(ratio_t[id(rec)]):
            if r is not None and r < 1:
                insights.append(f"RISIKO KEKURANGAN (Target) - {rec.label}: pada periode {period_labels[w]}, balance hanya {r*100:.0f}% dari target periode {period_labels[w+1]}.")
    for cabang, series in occ_t.items():
        for w, v in enumerate(series):
            if v is not None and v > 1:
                insights.append(f"RISIKO OVER KAPASITAS GUDANG (Target) - Cabang {cabang} pada periode {period_labels[w]}: occupancy {v*100:.0f}% dari kapasitas gudang.")
    for rec in records:
        yt = bal_t[id(rec)]
        delta = yt[-1] - yt[0]
        pct = (delta / abs(yt[0]) * 100) if yt[0] else None
        arah = "naik" if delta > 0 else ("turun" if delta < 0 else "stabil")
        if pct is not None and abs(pct) >= 30:
            insights.append(f"TREN - {rec.label} (Target): balance {arah} signifikan dari {yt[0]:.1f} ({period_labels[0]}) menjadi {yt[-1]:.1f} ({period_labels[-1]}), perubahan {pct:+.0f}%. " + ("Cek risiko overstock." if delta > 0 else "Cek risiko kekurangan stok."))
    if not insights:
        insights.append("Tidak ditemukan indikasi risiko kekurangan stok atau over-kapasitas gudang pada rentang periode yang dianalisis.")
    return insights[:50]

def build_html_report_string(charts_data, insights, period_labels, week_awal):
    imgs_html = "".join(f'<div class="chart"><h3>{title}</h3><img src="data:image/png;base64,{b64}" alt="{title}"/></div>' for title, b64 in charts_data)
    insight_items = "".join(f"<li>{i}</li>" for i in insights)
    return f"""<!DOCTYPE html>
<html lang="id">
<head>
<meta charset="utf-8"/>
<title>MRP Analysis Report</title>
<style>
  body {{ font-family: system-ui, -apple-system, "Segoe UI", sans-serif; background:#f9f9f7; color:#0b0b0b; margin:0; padding:32px; }}
  h1 {{ font-size: 22px; margin-bottom:4px; }}
  .sub {{ color:#52514e; margin-bottom:28px; }}
  .chart {{ background:#fcfcfb; border:1px solid #e1e0d9; border-radius:10px; padding:16px; margin-bottom:24px; }}
  .chart h3 {{ margin-top:0; font-size:15px; }}
  .chart img {{ max-width:100%; display:block; }}
  .insights {{ background:#fcfcfb; border:1px solid #e1e0d9; border-radius:10px; padding:20px 28px; }}
  .insights li {{ margin-bottom:10px; line-height:1.5; }}
  .grid {{ display:grid; grid-template-columns: 1fr 1fr; gap:20px; }}
  @media (max-width: 900px) {{ .grid {{ grid-template-columns: 1fr; }} }}
</style>
</head>
<body>
  <h1>MRP Balance &amp; Occupancy - Analysis Report</h1>
  <div class="sub">Periode: {period_labels[0]} - {period_labels[-1]} (Week Awal = {week_awal})</div>
  <div class="grid">{imgs_html}</div>
  <h2>Insight &amp; Rekomendasi</h2>
  <div class="insights"><ul>{insight_items}</ul></div>
</body>
</html>"""

def build_to_recommendations(records, bal_t, period_labels, n_weeks) -> list:
    """
    Rekomendasi TO (Transfer Order): sandingkan Shortage & Overstock Alerts
    pada kombinasi (Minggu, Grup, Category) yang SAMA PERSIS -- CABANG SENGAJA
    TIDAK ikut jadi match key, karena tujuan TO justru memindah stok ANTAR
    CABANG. Contoh nyata: Makassar shortage 6 unit "WARDROBE - OKELO S03FM" di
    NOV-4, sementara Manado overstock 5 unit kombinasi Grup+Category yang SAMA
    PERSIS di minggu yang SAMA -- itu kandidat TO yang valid walau cabang-nya
    beda (mewajibkan cabang yang sama juga akan membuat tabel ini nyaris selalu
    kosong: satu baris data tidak mungkin sekaligus minus DAN plus di minggu
    yang sama, jadi match hanya bisa kejadian lintas-cabang).

    Langkah:
    1. AGREGASI (setara GROUP BY + SUM di SQL) -- untuk tiap record/SKU per
       minggu, balance `bt[w]` dijumlah ke key (week, grup, category), DIPECAH
       per cabang di masing-masing sisi: bt[w] < 0 -> shortage[cabang] += |bt[w]|,
       bt[w] > 0 -> overstock[cabang] += bt[w]. Satu pass O(records x weeks),
       reuse `bal_t` yang sudah dihitung shortage_alerts/overstock_alerts di
       atas -- tidak ada perhitungan ganda.
    2. INNER JOIN kedua sisi pada key (week, grup, category): key yang cuma
       punya shortage TANPA overstock (atau sebaliknya) dibuang. Untuk key yang
       lolos, SETIAP cabang yang shortage dipasangkan dengan TOTAL overstock
       seluruh cabang lain pada key yang sama (satu baris output per cabang
       yang shortage) -- daftar cabang sumber overstock-nya disertakan di
       `overstock_sources` (diurutkan dari nilai terbesar) supaya rekomendasi
       tetap bisa ditindaklanjuti (tahu harus tarik dari cabang mana).
    3. Rekomendasi TO = MIN(shortage_value, overstock_value): jumlah yang
       secara fisik BISA dipindah dibatasi oleh sisi yang lebih kecil (tidak
       mungkin transfer melebihi defisit yang dibutuhkan penerima, atau
       melebihi total kelebihan stok yang tersedia di jaringan cabang lain).
       CATATAN: ini penjumlahan sederhana (bukan solver alokasi optimal) --
       kalau ada BEBERAPA cabang yang sama-sama shortage pada key yang sama,
       masing-masing dibandingkan ke total overstock yang sama (tidak saling
       mengurangi "jatah" satu sama lain).

    Setara SQL:
        WITH shortage_agg AS (
            SELECT week, grup, category, cabang, SUM(deficit) AS shortage_value
            FROM shortage_alerts GROUP BY week, grup, category, cabang
        ), overstock_agg AS (
            SELECT week, grup, category, SUM(excess) AS overstock_value
            FROM overstock_alerts GROUP BY week, grup, category
        )
        SELECT s.week, s.cabang, s.grup, s.category, s.shortage_value,
               o.overstock_value, LEAST(s.shortage_value, o.overstock_value) AS recommended_to
        FROM shortage_agg s
        INNER JOIN overstock_agg o
          ON s.week = o.week AND s.grup = o.grup AND s.category = o.category
        ORDER BY recommended_to DESC;
    """
    agg = {}  # (week, grup, category) -> {"shortage": {cabang: value}, "overstock": {cabang: value}}
    for rec in records:
        bt = bal_t[id(rec)]
        for w in range(n_weeks):
            val = bt[w]
            if val == 0:
                continue
            key = (period_labels[w], rec.grup, rec.category)
            entry = agg.setdefault(key, {"shortage": {}, "overstock": {}})
            side = entry["shortage"] if val < 0 else entry["overstock"]
            side[rec.cabang] = side.get(rec.cabang, 0.0) + (abs(val) if val < 0 else val)

    recommendations = []
    for (week, grup, category), sides in agg.items():
        shortage_by_cabang = sides["shortage"]
        overstock_by_cabang = sides["overstock"]
        total_overstock = sum(overstock_by_cabang.values())
        # INNER JOIN: hanya key yang punya KEDUA sisi (shortage & overstock)
        if not shortage_by_cabang or total_overstock <= 0:
            continue

        overstock_sources = sorted(
            ({"cabang": c, "value": round(v, 2)} for c, v in overstock_by_cabang.items()),
            key=lambda x: x["value"], reverse=True
        )
        for cabang, shortage_value in shortage_by_cabang.items():
            recommendations.append({
                "week": week,
                "cabang": cabang,
                "grup": grup,
                "category": category,
                "shortage_value": round(shortage_value, 2),
                "overstock_value": round(total_overstock, 2),
                "recommended_to": round(min(shortage_value, total_overstock), 2),
                "overstock_sources": overstock_sources,
            })

    recommendations.sort(key=lambda r: r["recommended_to"], reverse=True)
    return recommendations


def flag_occupancy_anomalies(daily_data: list) -> None:
    """Flag statistically unusual (cabang, week) rows via Isolation Forest.

    Additive-only: mutates each row in `daily_data` in place by adding an
    `is_anomaly` boolean, and never changes occupancy_pct/total_on_hand or
    any other existing field. This is a data-quality signal on TOP OF the
    deterministic running-balance calculation above, not a replacement for
    any part of it - a row can be "correct" per the formula and still get
    flagged here because it's an outlier versus the rest of the same upload
    (e.g. a fat-fingered On Hand or Target figure).

    Runs once per upload, on whatever rows this batch produced - there is no
    minimum-row check beyond what IsolationForest itself needs, since a
    small upload is exactly common for this module (a handful of cabang).
    """
    for row in daily_data:
        row["is_anomaly"] = False

    if len(daily_data) < 10:
        # Too few rows for "unusual vs. the rest of this batch" to mean
        # anything - every row would look equally unique.
        return

    features = np.array([
        [d["occupancy_pct"], d["total_on_hand"]] for d in daily_data
    ], dtype=float)

    try:
        model = IsolationForest(contamination=0.05, random_state=42, n_estimators=100)
        preds = model.fit_predict(features)  # -1 = anomaly, 1 = normal
    except Exception:
        return

    for row, pred in zip(daily_data, preds):
        row["is_anomaly"] = bool(pred == -1)


def project_occupancy_forward(occ_t: dict, week_awal: int, n_weeks: int, steps: int = 3) -> dict:
    """Project each cabang's occupancy ratio `steps` weeks beyond the
    uploaded horizon, using the same genuine Holt-Winters fit used by the
    Demand Forecasting module (see services/forecast_engine.py) - not a new,
    separately-invented method.

    Purely additive: this never feeds back into daily_data, shortage/overstock
    alerts, TO recommendations, or the generated Excel workbook - it is an
    extra, clearly-separate "what's coming next" signal for the dashboard.
    Falls back silently (empty projection for that cabang) if the fit fails
    or history is too short, same fallback behavior as forecast_engine.
    """
    projection = {}
    for cabang, series in occ_t.items():
        clean_series = [v if v is not None else 0.0 for v in series]
        if len(clean_series) < 4:
            continue
        try:
            preds = _holt_winters_forecast(clean_series, steps)
        except Exception:
            continue
        projection[str(cabang)] = [round(max(0.0, p) * 100, 2) for p in preds]

    future_labels = [period_label(week_awal + n_weeks - 1 + i) for i in range(1, steps + 1)] if projection else []
    return {"series_by_cabang": projection, "period_labels": future_labels}


def calculate_mrp_occupancy_from_bytes(file_bytes: bytes) -> dict:
    # read_only=True: pass ini cuma pernah memanggil iter_rows()/sheetnames pada
    # wb_data (nilai terhitung, bukan formula) -- tidak pernah menulis atau
    # mengakses style, jadi read_only aman dipakai dan JAUH lebih cepat/hemat
    # memori untuk file besar (openpyxl tidak membangun model penuh per-sel
    # dengan style di RAM). wb_form di bawah (dipakai untuk MENULIS sheet hasil)
    # sengaja TETAP di-load ulang non-read_only karena butuh mempertahankan
    # formula/format asli sheet Raw/WH dan harus bisa di-save().
    wb_data = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    if "Raw" not in wb_data.sheetnames or "WH" not in wb_data.sheetnames:
        raise ValueError("File Excel harus memiliki sheet 'Raw' dan 'WH' untuk pemrosesan MRP.")

    ws_raw, ws_wh = wb_data["Raw"], wb_data["WH"]
    raw_cols = resolve_raw_columns(ws_raw)
    n_weeks = detect_week_count(ws_raw, raw_cols)
    week_awal = read_week_awal(ws_wh, default=1)
    period_labels = build_period_labels(week_awal, n_weeks)

    records = read_raw_records(ws_raw, n_weeks, raw_cols)
    wh_capacity = read_wh_capacity(ws_wh)

    # B.4: Sheet Harga Container
    harga_dict = {}
    if "Harga Container" in wb_data.sheetnames:
        ws_harga = wb_data["Harga Container"]
        for row in ws_harga.iter_rows(min_row=2, values_only=True):
            if len(row) >= 4:
                cab = str(row[1]).strip() if row[1] else ""
                grp = str(row[2]).strip() if row[2] else ""
                hrg = _safe_float(row[3])
                if cab and grp:
                    harga_dict[(cab, grp)] = hrg

    # Sheet CBM & Harga Product (key: Cabang+Category) -- dipakai untuk konversi
    # Container -> QTY -> Rupiah, sama seperti dsp_calculator_conversion.py.
    def _read_lookup_sheet(sheet_name):
        lookup = {}
        if sheet_name not in wb_data.sheetnames:
            return lookup
        ws = wb_data[sheet_name]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if len(row) >= 4:
                cab = str(row[1]).strip() if row[1] else ""
                cat = str(row[2]).strip() if row[2] else ""
                val = _safe_float(row[3])
                if cab and cat:
                    lookup[(cab, cat)] = val
        return lookup

    cbm_dict = _read_lookup_sheet("CBM")
    harga_product_dict = _read_lookup_sheet("Harga Product")

    bal_t, ratio_t = {}, {}
    for rec in records:
        bt = compute_balance_series(rec, n_weeks)
        bal_t[id(rec)] = bt
        ratio_t[id(rec)] = compute_ratio_series(bt, rec, n_weeks)

    occ_t = compute_occupancy(records, bal_t, wh_capacity)

    # Konversi Container -> QTY -> Rupiah per record (per Cabang+Grup+Category),
    # lalu diagregasi per cabang untuk ditampilkan di dashboard, dan disimpan
    # per-record untuk dituliskan ke sheet "1. Running Balance" hasil Excel.
    qty_rupiah_by_record = {}
    qty_series_by_branch, value_series_by_branch = {}, {}
    for rec in records:
        qty_series, rupiah_series = compute_qty_rupiah_series(
            rec.cabang, rec.category, bal_t[id(rec)], n_weeks, cbm_dict, harga_product_dict
        )
        qty_rupiah_by_record[id(rec)] = (qty_series, rupiah_series)

        qty_acc = qty_series_by_branch.setdefault(rec.cabang, [0.0] * n_weeks)
        value_acc = value_series_by_branch.setdefault(rec.cabang, [0.0] * n_weeks)
        for w in range(n_weeks):
            if qty_series[w] is not None:
                qty_acc[w] += qty_series[w]
            if rupiah_series[w] is not None:
                value_acc[w] += rupiah_series[w]

    # Fitur 3: MOS berbasis Value (AA = Target QTY x Harga Product x 4, SUM per
    # cabang per minggu, lalu Value(Rp) / SUM(AA)) -- lihat compute_mos_value_series_by_branch.
    mos_value_series_by_branch = compute_mos_value_series_by_branch(records, n_weeks, cbm_dict, harga_product_dict, value_series_by_branch)

    # Analisa Nilai Inventori -- level Regional (Group By kolom Region) &
    # Nasional (satu key NATIONAL_KEY, jadi total gabungan semua cabang/region).
    # Reuse qty_rupiah_by_record yang sama dengan level Cabang di atas supaya
    # SUM(Regional) == SUM(Cabang) == Nasional persis cocok (lihat
    # aggregate_qty_value_by_group/compute_mos_value_series_by_group).
    qty_series_by_region, value_series_by_region = aggregate_qty_value_by_group(
        records, n_weeks, qty_rupiah_by_record, group_key_fn=lambda rec: rec.region)
    mos_value_series_by_region = compute_mos_value_series_by_group(
        records, n_weeks, cbm_dict, harga_product_dict, value_series_by_region,
        group_key_fn=lambda rec: rec.region)

    qty_series_by_national, value_series_by_national = aggregate_qty_value_by_group(
        records, n_weeks, qty_rupiah_by_record, group_key_fn=lambda rec: NATIONAL_KEY)
    mos_value_series_by_national = compute_mos_value_series_by_group(
        records, n_weeks, cbm_dict, harga_product_dict, value_series_by_national,
        group_key_fn=lambda rec: NATIONAL_KEY)

    regions_list = sorted({rec.region for rec in records})

    # B.5: Implementasi MOS
    # TAHAP 5.a: Agregasi Hasil Target
    agg_bal_t = {}
    agg_target = {} # TAHAP 5.c
    for rec in records:
        key = (rec.cabang, rec.grup)
        if key not in agg_bal_t:
            agg_bal_t[key] = [0.0] * n_weeks
            agg_target[key] = [0.0] * n_weeks
        bt = bal_t[id(rec)]
        for w in range(n_weeks):
            agg_bal_t[key][w] += bt[w]
            agg_target[key][w] += rec.target[w]

    print("AGG_BAL_T:", agg_bal_t)
    mos_rows = []
    for (cab, grp), bals in agg_bal_t.items():
        harga = harga_dict.get((cab, grp), 0.0)
        targs = agg_target[(cab, grp)]
        for w in range(n_weeks):
            val_perhitungan = bals[w]
            # TAHAP 5.b: Perkalian Nilai Inventory
            nilai_inventory = val_perhitungan * harga
            
            # TAHAP 5.d: Kalkulasi Final MOS
            # MOS = Balance / (Target * 4). Fallback ke 0 saat (Target * 4) = 0
            # (Target = 0), supaya pembagian oleh nol tidak melempar exception.
            val_target = targs[w]
            target_x4 = val_target * 4
            mos = (val_perhitungan / target_x4) if target_x4 != 0 else 0.0
            
            mos_rows.append({
                "Cabang": cab,
                "Grup": grp,
                "Week": period_labels[w],
                "Balance": val_perhitungan,
                "Target": val_target,
                "Harga": harga,
                "Value_per_Week": nilai_inventory,
                "MOS": mos
            })
            
    import pandas as pd
    import numpy as np
    df_mos = pd.DataFrame(mos_rows)
    if not df_mos.empty:
        df_mos.replace([np.inf, -np.inf], 0, inplace=True)
        df_mos.fillna(0, inplace=True)

    b64_bal_t = plot_balance_chart_b64(records, bal_t, period_labels, "Tren Balance (Group) - Skenario Target")
    b64_occ_t = plot_occupancy_chart_b64(occ_t, period_labels, "Occupancy Gudang - Skenario Target")

    insights = generate_insights(records, bal_t, ratio_t, occ_t, period_labels, n_weeks)
    html_report = build_html_report_string([
        ("Balance (Group) - Skenario Target", b64_bal_t),
        ("Occupancy - Skenario Target", b64_occ_t),
    ], insights, period_labels, week_awal)

    wb_form = load_workbook(io.BytesIO(file_bytes), data_only=False)
    wb_form, _, _, _ = generate_excel_workbook(wb_form, records=records, qty_rupiah_by_record=qty_rupiah_by_record)
            
    out_buf = io.BytesIO()
    wb_form.save(out_buf)
    out_buf.seek(0)
    excel_base64 = base64.b64encode(out_buf.read()).decode("ascii")

    # Breakdown per (Cabang, Week) untuk tooltip chart Occupancy (Stock Awal,
    # Vessel inbound, TO, Target penjualan, dan komposisi per Grup) -- dihitung
    # dari field yang sama yang sudah dipakai shortage/overstock di atas
    # (rec.onhand/.vessel/.to/.target dan running balance bt), cuma diagregasi
    # ke level cabang+week di sini. `grup` dikirim TANPA dipotong top-N --
    # sorting & slice top-3 sengaja dilakukan di frontend (lihat OccupancyChart).
    breakdown_by_cabang_week = {}
    for rec in records:
        bt = bal_t[id(rec)]
        for w in range(n_weeks):
            key = (rec.cabang, w)
            acc = breakdown_by_cabang_week.setdefault(key, {
                "stock_awal": 0.0, "vessel_in": 0.0, "to": 0.0, "target": 0.0, "grup": {}
            })
            stock_awal_w = rec.onhand if w == 0 else bt[w - 1]
            acc["stock_awal"] += stock_awal_w
            acc["vessel_in"] += rec.vessel[w]
            acc["to"] += rec.to[w]
            acc["target"] += rec.target[w]
            # Kuantitas per grup dalam Container -- pakai balance akhir minggu ini
            # (max 0, sama seperti compute_occupancy: shortage/minus tidak
            # membebaskan ruang fisik gudang), bukan target/demand.
            acc["grup"][rec.grup] = acc["grup"].get(rec.grup, 0.0) + max(0.0, bt[w])

    daily_data = []
    for cabang, cap_val in wh_capacity:
        series_occ = occ_t.get(cabang, [0.0] * n_weeks)
        for w in range(n_weeks):
            occ_pct = series_occ[w] if series_occ[w] is not None else 0.0
            tot_bal = occ_pct * cap_val if cap_val else 0.0
            b = breakdown_by_cabang_week.get((cabang, w), {})
            # Cap 15 grup TERBESAR per baris -- frontend cuma butuh top-3, jadi
            # angka ini cuma jaring pengaman ukuran payload (bukan nilai yang
            # dianggap "final") untuk cabang dengan jumlah grup yang sangat
            # banyak, sama motivasinya dengan INVENTORY_VALUE_ROWS_BYTE_BUDGET
            # di routers/occupancy.py.
            grup_items = sorted(b.get("grup", {}).items(), key=lambda kv: kv[1], reverse=True)[:15]
            daily_data.append({
                "cabang": str(cabang),
                "date": period_labels[w],
                "total_on_hand": round(tot_bal, 2),
                "capacity": round(cap_val, 2),
                "occupancy_pct": round(occ_pct * 100, 2),
                "is_shortage": tot_bal < 0,
                "breakdown": {
                    "stock_awal": round(b.get("stock_awal", 0.0), 2),
                    "vessel_in": round(b.get("vessel_in", 0.0), 2),
                    "to": round(b.get("to", 0.0), 2),
                    "target_penjualan": round(b.get("target", 0.0), 2),
                    "grup": [{"nama": g, "qty": round(q, 2)} for g, q in grup_items],
                },
            })

    # Additive-only data-quality pass - see flag_occupancy_anomalies docstring.
    # Adds `is_anomaly` to each row; never changes occupancy_pct/total_on_hand.
    flag_occupancy_anomalies(daily_data)

    shortage_alerts = []
    # Overstock = perhitungan SAMA PERSIS seperti Shortage di atas, cuma arah
    # tandanya dibalik -- tidak ada threshold/pengali tambahan (coverage-weeks,
    # target minggu depan, dll sudah dilepas atas permintaan langsung supaya
    # logikanya benar-benar cermin 1:1 dengan Shortage):
    #   Shortage  : bt[w] < 0  -> deficit = abs(bt[w])
    #   Overstock : bt[w] > 0  -> excess  = bt[w]
    overstock_alerts = []
    for rec in records:
        bt = bal_t[id(rec)]
        for w in range(n_weeks):
            if bt[w] < 0:
                shortage_alerts.append({
                    "cabang": rec.cabang,
                    "category": f"{rec.grup} - {rec.category}",
                    "date": period_labels[w],
                    "deficit": round(abs(bt[w]), 2)
                })
            elif bt[w] > 0:
                overstock_alerts.append({
                    "cabang": rec.cabang,
                    "category": f"{rec.grup} - {rec.category}",
                    "date": period_labels[w],
                    "excess": round(bt[w], 2)
                })

    # Rekomendasi TO (Transfer Order): sandingkan shortage & overstock pada
    # kombinasi (Minggu, Cabang, Grup, Category) yang sama persis -- lihat
    # build_to_recommendations di atas untuk detail langkah & versi SQL-nya.
    to_recommendations = build_to_recommendations(records, bal_t, period_labels, n_weeks)

    # Sheet 1 "Analisa Nilai Inventori": row per (Cabang, Grup, Category, Week)
    # dengan alur Container (balance) -> QTY (CBM) -> Value, memakai deret yang
    # SAMA dengan yang sudah dituliskan ke sheet "1. Running Balance" (lihat
    # append_qty_rupiah_blocks) -- di sini cuma diserialisasi juga ke JSON
    # supaya frontend bisa menyusun ulang sheet ini sesuai filter aktif di UI.
    inventory_value_rows = []
    for rec in records:
        bt = bal_t[id(rec)]
        qty_series, rupiah_series = qty_rupiah_by_record[id(rec)]
        harga_satuan = harga_product_dict.get((rec.cabang, rec.category))
        for w in range(n_weeks):
            inventory_value_rows.append({
                "region": rec.region,
                "cabang": rec.cabang,
                "grup": rec.grup,
                "category": rec.category,
                "week": period_labels[w],
                "balance_container": round(bt[w], 2),
                "qty": round(qty_series[w], 2) if qty_series[w] is not None else None,
                "harga_satuan": harga_satuan,
                "value": round(rupiah_series[w], 2) if rupiah_series[w] is not None else None,
            })

    avg_occ = sum(d["occupancy_pct"] for d in daily_data) / len(daily_data) if daily_data else 0
    max_occ = max(d["occupancy_pct"] for d in daily_data) if daily_data else 0
    anomaly_count = sum(1 for d in daily_data if d.get("is_anomaly"))

    # Predictive occupancy - projects occ_t (already computed above for
    # daily_data/charts) `steps` weeks past the uploaded horizon. Additive
    # only: see project_occupancy_forward docstring.
    occupancy_projection = project_occupancy_forward(occ_t, week_awal, n_weeks)

    # NOTE: `mrp_results` used to be duplicated verbatim into a legacy
    # `ddmrp_results` key (a leftover from a module rename), and each of the
    # two balance/occupancy chart PNGs was embedded both standalone AND baked
    # into `html_report`. The frontend only ever reads `mrp_results` (falling
    # back to `ddmrp_results` only when the former is absent, which never
    # happens here), and never reads the standalone `charts` field. That
    # quadrupled the JSON body for no functional benefit and was the direct
    # cause of Vercel's 4.5MB FUNCTION_PAYLOAD_TOO_LARGE errors on this
    # endpoint. Keep a single copy of everything.
    return {
        "processed_at": datetime.now().isoformat(),
        "daily_data": daily_data,
        "shortage_alerts": shortage_alerts,
        "overstock_alerts": overstock_alerts,
        "to_recommendations": to_recommendations,
        "inventory_value_rows": inventory_value_rows,
        "kpi_summary": {
            "avg_occupancy": round(avg_occ, 1),
            "max_occupancy": round(max_occ, 1),
            "categories_at_risk": len(shortage_alerts),
            "overstock_count": len(overstock_alerts),
            "to_recommendation_count": len(to_recommendations),
            "anomaly_count": anomaly_count,
        },
        "mos_data": df_mos.to_dict('records') if not df_mos.empty else [],
        "over_occupancy_insights": insights,
        "inventory_analysis": None,
        "mrp_results": {
            "week_awal": week_awal,
            "period_labels": period_labels,
            "html_report": html_report,
            "excel_base64": excel_base64,
            "insights_list": insights,
            "occupancy_series_target": {str(k): [round(v * 100, 2) if v is not None else 0 for v in val] for k, val in occ_t.items()},
            # Predictive occupancy (Holt-Winters, same fit used by Demand
            # Forecasting) - projects `steps` weeks past the uploaded horizon.
            # Additive-only: see project_occupancy_forward docstring.
            "occupancy_projection": occupancy_projection,
            # Konversi Container -> QTY -> Rupiah (lookup CBM & Harga Product per
            # Cabang+Category), diagregasi per cabang per minggu -- dipakai untuk
            # tabel "Analisa Nilai Inventori (QTY & Value)" di dashboard.
            "qty_series_by_branch": {str(k): [round(v, 2) for v in val] for k, val in qty_series_by_branch.items()},
            "value_series_by_branch": {str(k): [round(v, 2) for v in val] for k, val in value_series_by_branch.items()},
            # Fitur 3: MOS berbasis Value, per cabang per minggu -- ditampilkan
            # tepat di bawah baris "Value (Rp)" pada tabel Analisa Nilai Inventori.
            "mos_value_series_by_branch": {str(k): v for k, v in mos_value_series_by_branch.items()},
            # Analisa Nilai Inventori -- level Regional & Nasional (lihat
            # aggregate_qty_value_by_group/compute_mos_value_series_by_group).
            # Field baru, additive-only: tidak mengganti/menghapus field di atas.
            "regions_list": regions_list,
            "qty_series_by_region": {str(k): [round(v, 2) for v in val] for k, val in qty_series_by_region.items()},
            "value_series_by_region": {str(k): [round(v, 2) for v in val] for k, val in value_series_by_region.items()},
            "mos_value_series_by_region": {str(k): v for k, v in mos_value_series_by_region.items()},
            "national_summary": {
                "qty_series": [round(v, 2) for v in qty_series_by_national.get(NATIONAL_KEY, [0.0] * n_weeks)],
                "value_series": [round(v, 2) for v in value_series_by_national.get(NATIONAL_KEY, [0.0] * n_weeks)],
                "mos_value_series": mos_value_series_by_national.get(NATIONAL_KEY, [0.0] * n_weeks),
            },
        },
    }


def generate_mrp_template_bytes() -> bytes:
    """
    Menghasilkan file Excel template resmi MRP Occupancy & Inventory Projector
    lengkap dengan sheet 'Raw' (No, Region, Cabang, Grup, Category, On Hand,
    lalu blok 4 kolom per minggu [TO, Vessel, Buffer, Target]) dan sheet 'WH'
    (Kapasitas & Week Awal). Kolom "Region" di posisi B ini yang dideteksi
    resolve_raw_columns() -- lihat docstring fungsi itu untuk kompatibilitas
    mundur dengan file yang masih pakai template lama (tanpa kolom Region).
    """
    wb = Workbook()
    ws_raw = wb.active
    ws_raw.title = "Raw"
    ws_wh = wb.create_sheet("WH")
    ws_harga = wb.create_sheet("Harga Container")
    ws_cbm = wb.create_sheet("CBM")
    ws_harga_product = wb.create_sheet("Harga Product")

    n_weeks = 6
    col_start_week = 7  # Kolom G (1-indexed = 7) -- geser 1 kolom dari sebelumnya karena Region disisipkan di B

    for c in range(1, col_start_week):
        style_header_cell(ws_raw, 1, c, "Judul", FILL_JUDUL)
    for w in range(n_weeks):
        base_c = col_start_week + (w * 4)
        for sub_c in range(4):
            style_header_cell(ws_raw, 1, base_c + sub_c, f"Week {w + 1}", FILL_PERHITUNGAN if w % 2 == 0 else FILL_RATIO)

    fixed_headers = ["No", "Region", "Cabang", "Grup", "Category", "On Hand"]
    for i, h in enumerate(fixed_headers, start=1):
        style_header_cell(ws_raw, 2, i, h, FILL_HEADER2)

    for w in range(n_weeks):
        base_c = col_start_week + (w * 4)
        style_header_cell(ws_raw, 2, base_c, "TO", FILL_HEADER2)
        style_header_cell(ws_raw, 2, base_c + 1, "Vessel", FILL_HEADER2)
        style_header_cell(ws_raw, 2, base_c + 2, "Buffer", FILL_HEADER2)
        style_header_cell(ws_raw, 2, base_c + 3, "Target", FILL_HEADER2)

    sample_raw = [
        [1, "Jawa", "DC Jakarta", "FMCG", "Beverages", 15000],
        [2, "Jawa", "DC Surabaya", "Electronics", "Gadgets", 8500],
        [3, "Sumatera", "DC Medan", "Apparel", "Fashion Casual", 4200],
        [4, "Sulawesi", "DC Makassar", "Automotive", "Spareparts", 6300]
    ]
    sample_weeks_data = [
        [[1200, 2000, 3000, 3200], [1500, 1000, 3200, 3500], [1000, 2500, 3100, 3400], [2000, 1500, 3500, 3800], [1800, 2000, 3300, 3600], [1500, 1800, 3400, 3700]],
        [[500, 800, 1500, 1400], [600, 900, 1600, 1500], [700, 500, 1700, 1600], [800, 1000, 1800, 1700], [600, 700, 1600, 1500], [500, 800, 1500, 1400]],
        [[200, 300, 600, 700], [300, 400, 700, 800], [250, 350, 650, 750], [400, 500, 800, 900], [300, 400, 700, 800], [200, 300, 600, 700]],
        [[400, 600, 1100, 1200], [500, 700, 1200, 1300], [450, 650, 1150, 1250], [600, 800, 1300, 1400], [500, 700, 1200, 1300], [400, 600, 1100, 1200]],
    ]

    for row_idx, row_base in enumerate(sample_raw, start=3):
        for col_idx, val in enumerate(row_base, start=1):
            cell = ws_raw.cell(row=row_idx, column=col_idx, value=val)
            cell.border = BORDER
        weeks_vals = sample_weeks_data[row_idx - 3]
        for w in range(n_weeks):
            base_c = col_start_week + (w * 4)
            for sub_c, v in enumerate(weeks_vals[w]):
                cell = ws_raw.cell(row=row_idx, column=base_c + sub_c, value=v)
                cell.border = BORDER

    ws_raw.column_dimensions["A"].width = 6
    ws_raw.column_dimensions["B"].width = 14
    ws_raw.column_dimensions["C"].width = 15
    ws_raw.column_dimensions["D"].width = 14
    ws_raw.column_dimensions["E"].width = 16
    ws_raw.column_dimensions["F"].width = 12
    for c in range(col_start_week, col_start_week + (n_weeks * 4)):
        ws_raw.column_dimensions[get_column_letter(c)].width = 11
    ws_raw.freeze_panes = "G3"

    wh_headers = ["No", "Cabang", "Kapasitas Existing", "Tambahan", "Total Kapasitas"]
    for i, h in enumerate(wh_headers, start=1):
        style_header_cell(ws_wh, 1, i, h, FILL_HEADER2)

    ws_wh.cell(row=1, column=7, value="Week Awal").font = BOLD
    ws_wh.cell(row=1, column=7).fill = FILL_JUDUL
    ws_wh.cell(row=1, column=7).border = BORDER
    ws_wh.cell(row=1, column=8, value=1).font = BOLD
    ws_wh.cell(row=1, column=8).fill = FILL_RATIO
    ws_wh.cell(row=1, column=8).border = BORDER
    ws_wh.cell(row=1, column=8).alignment = CENTER

    sample_wh = [
        [1, "DC Jakarta", 25000, 5000],
        [2, "DC Surabaya", 15000, 2000],
        [3, "DC Medan", 8000, 1000],
        [4, "DC Makassar", 12000, 1500],
    ]
    for row_idx, row_data in enumerate(sample_wh, start=2):
        for col_idx, val in enumerate(row_data, start=1):
            cell = ws_wh.cell(row=row_idx, column=col_idx, value=val)
            cell.border = BORDER
        tot_cell = ws_wh.cell(row=row_idx, column=5, value=f"=C{row_idx}+D{row_idx}")
        tot_cell.border = BORDER
        tot_cell.font = BOLD

    ws_wh.column_dimensions["A"].width = 6
    ws_wh.column_dimensions["B"].width = 16
    ws_wh.column_dimensions["C"].width = 18
    ws_wh.column_dimensions["D"].width = 14
    ws_wh.column_dimensions["E"].width = 18
    ws_wh.column_dimensions["G"].width = 14
    ws_wh.column_dimensions["H"].width = 10

    # B.4: Template Harga Container
    harga_headers = ["No", "Cabang", "Grup", "Harga"]
    for i, h in enumerate(harga_headers, start=1):
        style_header_cell(ws_harga, 1, i, h, FILL_HEADER2)
    
    sample_harga = [
        [1, "DC Jakarta", "FMCG", 15000000],
        [2, "DC Surabaya", "Electronics", 25000000],
        [3, "DC Medan", "Apparel", 12000000],
        [4, "DC Makassar", "Automotive", 18000000]
    ]
    for row_idx, row_data in enumerate(sample_harga, start=2):
        for col_idx, val in enumerate(row_data, start=1):
            cell = ws_harga.cell(row=row_idx, column=col_idx, value=val)
            cell.border = BORDER
            
    ws_harga.column_dimensions["B"].width = 15
    ws_harga.column_dimensions["C"].width = 15
    ws_harga.column_dimensions["D"].width = 18

    # B.5: Template CBM (lookup volume per unit, key = Cabang + Category --
    # skema ini disamakan persis dengan build_lookup_cbm() di
    # dsp_calculator_conversion.py: kolom "Nilai" dipakai sebagai lookup value
    # saat konversi Container -> QTY, Cabang+Category adalah kunci merge-nya)
    cbm_headers = ["No", "Cabang", "Category", "Nilai"]
    for i, h in enumerate(cbm_headers, start=1):
        style_header_cell(ws_cbm, 1, i, h, FILL_HEADER2)

    sample_cbm = [
        [1, "DC Jakarta", "Beverages", 0.015],
        [2, "DC Surabaya", "Gadgets", 0.008],
        [3, "DC Medan", "Fashion Casual", 0.004],
        [4, "DC Makassar", "Spareparts", 0.020],
    ]
    for row_idx, row_data in enumerate(sample_cbm, start=2):
        for col_idx, val in enumerate(row_data, start=1):
            cell = ws_cbm.cell(row=row_idx, column=col_idx, value=val)
            cell.border = BORDER

    ws_cbm.column_dimensions["A"].width = 6
    ws_cbm.column_dimensions["B"].width = 15
    ws_cbm.column_dimensions["C"].width = 16
    ws_cbm.column_dimensions["D"].width = 14

    # B.6: Template Harga Product (lookup harga per unit, key = Cabang + Category --
    # skema ini disamakan persis dengan build_lookup_harga_product() di
    # dsp_calculator_conversion.py: kolom "Nilai" dipakai sebagai lookup value
    # saat konversi QTY -> Rupiah. Berbeda dari sheet "Harga Container" di atas
    # yang granularitasnya Cabang + Grup saja untuk kebutuhan sheet "3. Harga & MOS")
    harga_product_headers = ["No", "Cabang", "Category", "Nilai"]
    for i, h in enumerate(harga_product_headers, start=1):
        style_header_cell(ws_harga_product, 1, i, h, FILL_HEADER2)

    sample_harga_product = [
        [1, "DC Jakarta", "Beverages", 15000],
        [2, "DC Surabaya", "Gadgets", 25000000],
        [3, "DC Medan", "Fashion Casual", 120000],
        [4, "DC Makassar", "Spareparts", 850000],
    ]
    for row_idx, row_data in enumerate(sample_harga_product, start=2):
        for col_idx, val in enumerate(row_data, start=1):
            cell = ws_harga_product.cell(row=row_idx, column=col_idx, value=val)
            cell.border = BORDER

    ws_harga_product.column_dimensions["A"].width = 6
    ws_harga_product.column_dimensions["B"].width = 15
    ws_harga_product.column_dimensions["C"].width = 16
    ws_harga_product.column_dimensions["D"].width = 14

    out_buf = io.BytesIO()
    wb.save(out_buf)
    out_buf.seek(0)
    return out_buf.read()
