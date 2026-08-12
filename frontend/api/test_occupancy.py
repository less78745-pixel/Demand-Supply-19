import sys
sys.path.append('c:\\Users\\DELL\\Downloads\\python\\Vibes Coding\\wms-monorepo\\frontend\\api')
from services.occupancy_engine import calculate_mrp_occupancy_from_bytes, read_raw_records, detect_week_count
import openpyxl, io

wb = openpyxl.Workbook()
ws = wb.active; ws.title = 'Raw'
ws.append(['Cabang', 'Category', 'Grup', 'L/T', 'M', 'M+1', 'Avg', 'Target/Wk', 'Harga', 'M', 'M+1', 'Total', 'Week 1', 'Week 1', 'Week 1', 'Week 1', 'Week 2', 'Week 2', 'Week 2', 'Week 2'])
ws.append(['DC Jakarta', 'Cat1', 'Grp1', 1, 10, 10, 10, 10, 15000, 10, 10, 20, 15000, 1200, 2000, 3200, 14000, 1500, 2500, 3500])
ws_wh = wb.create_sheet('WH')
ws_wh.append(['Judul', 'Kapasitas', 'Week Awal'])
ws_wh.append(['DC Jakarta', 30000, 1])
ws_harga = wb.create_sheet('Harga Container')
ws_harga.append(['No', 'Cabang', 'Grup', 'Harga'])
ws_harga.append([1, 'DC Jakarta', 'Grp1', 15000])

out = io.BytesIO()
wb.save(out)
try:
    with open('c:\\Users\\DELL\\Downloads\\python\\Vibes Coding\\wms-monorepo\\frontend\\api\\services\\occupancy_engine.py', 'r') as f:
        src = f.read()
    if 'print("AGG_BAL_T:", agg_bal_t)' not in src:
        src = src.replace('mos_rows = []', 'print("AGG_BAL_T:", agg_bal_t)\n    mos_rows = []')
        with open('c:\\Users\\DELL\\Downloads\\python\\Vibes Coding\\wms-monorepo\\frontend\\api\\services\\occupancy_engine.py', 'w') as f:
            f.write(src)

    import importlib
    import frontend.api.services.occupancy_engine
    importlib.reload(frontend.api.services.occupancy_engine)
    res = frontend.api.services.occupancy_engine.calculate_mrp_occupancy_from_bytes(out.getvalue())
    print("MOS DATA:", res['mos_data'])
except Exception as e:
    import traceback
    traceback.print_exc()
